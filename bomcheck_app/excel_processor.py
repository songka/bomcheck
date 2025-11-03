# -*- coding: utf-8 -*-
from __future__ import annotations

from collections import defaultdict
import re
from math import isclose
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .binding_library import BindingChoice, BindingGroup, BindingLibrary
from .models import (
    BindingProjectResult,
    ExecutionResult,
    ImportantMaterialHit,
    MissingItem,
    ReplacementRecord,
    ReplacementSummary,
    RequirementGroupResult,
)
from .text_utils import normalize_text, normalized_variants


def normalize_part_no(value: str) -> str:
    return "".join(str(value).strip().upper().split())


BLACK_FILL = PatternFill(start_color="000000", end_color="000000", fill_type="solid")


def _is_black_fill(cell: Cell) -> bool:
    """判断单元格是否已被填充为黑色。"""
    fill = cell.fill
    if not fill or fill.fill_type != "solid":
        return False
    rgb = (fill.start_color.rgb or "").upper()
    return rgb in {"000000", "FF000000"}


def _row_has_non_black_value(row: Tuple[Cell, ...], ignore_idx: int) -> bool:
    """判断当前行是否存在未被涂黑的非空单元格（用于识别已追加的新料号）。"""
    for idx, cell in enumerate(row):
        if idx == ignore_idx:
            continue
        if cell.value not in (None, "") and not _is_black_fill(cell):
            return True
    return False


def _row_contains_part(row: Tuple[Cell, ...], ignore_idx: int, part_no: str) -> bool:
    if not part_no:
        return False
    normalized_target = normalize_part_no(part_no)
    for idx, cell in enumerate(row):
        if idx == ignore_idx:
            continue
        value = cell.value
        if not value:
            continue
        if normalize_part_no(str(value)) == normalized_target:
            return True
    return False


def format_quantity_cell(value):
    if value in (None, ""):
        return ""
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        try:
            number = float(value)
        except (TypeError, ValueError):
            return value
    if isclose(number, round(number), abs_tol=1e-6):
        return int(round(number))
    return round(number, 4)


def format_quantity_text(value) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        try:
            number = float(value)
        except (TypeError, ValueError):
            return str(value)
    if isclose(number, round(number), abs_tol=1e-6):
        return str(int(round(number)))
    text = f"{round(number, 4):g}"
    return text


_PART_NO_PATTERN = re.compile(r"^[A-Z0-9][A-Z0-9._/-]*$")


def _is_probable_part_number(value: str) -> bool:
    normalized = normalize_part_no(value)
    if not normalized:
        return False
    if not _PART_NO_PATTERN.fullmatch(normalized):
        return False
    if not any(ch.isdigit() for ch in normalized):
        return False
    if not any(ch.isalpha() for ch in normalized):
        return False
    return True


class ExcelProcessor:
    def __init__(self, config) -> None:
        self.config = config

    def execute(self, excel_path: Path, binding_library: BindingLibrary) -> ExecutionResult:
        wb = load_workbook(excel_path)

        for sheet_name in ("执行统计", "剩余物料"):
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]

        # 使用除结果汇总外的所有工作表参与业务处理
        data_sheets = [
            ws for ws in wb.worksheets if ws.title not in {"执行统计", "剩余物料"}
        ]

        debug_logs: List[str] = []

        replacement_summary, replacement_debug = self._apply_replacements(data_sheets)
        debug_logs.extend(replacement_debug)

        (
            part_quantities,
            part_desc,
            part_display,
            quantity_debug,
        ) = self._extract_part_quantities(data_sheets)
        debug_logs.extend(quantity_debug)

        # Apply replacements to aggregated data
        for record in replacement_summary.records:
            if not record.replacement_part_no:
                continue
            invalid_key = normalize_part_no(record.invalid_part_no)
            replacement_key = normalize_part_no(record.replacement_part_no)

            qty = part_quantities.pop(invalid_key, 0.0)
            part_desc.pop(invalid_key, None)
            part_display.pop(invalid_key, None)

            if qty:
                part_quantities[replacement_key] += qty
            part_display.setdefault(replacement_key, record.replacement_part_no)
            if record.replacement_desc:
                part_desc.setdefault(replacement_key, record.replacement_desc)

        available_inventory: Dict[str, float] = defaultdict(float, part_quantities)

        (
            binding_results,
            missing_items,
            used_part_numbers,
            binding_debug,
        ) = self._evaluate_binding_requirements(
            part_quantities,
            available_inventory,
            part_desc,
            part_display,
            binding_library,
        )
        debug_logs.extend(binding_debug)

        important_hits, important_part_numbers, important_debug = self._scan_important_materials(
            part_quantities,
            part_display,
            part_desc,
        )
        debug_logs.extend(important_debug)

        remainder_keys = (set(part_quantities.keys()) - used_part_numbers) | important_part_numbers

        self._write_result_sheets(
            wb,
            replacement_summary,
            binding_results,
            important_hits,
            missing_items,
            part_quantities,
            part_desc,
            part_display,
            remainder_keys,
            debug_logs,
        )

        wb.save(excel_path)

        return ExecutionResult(
            replacement_summary=replacement_summary,
            binding_results=binding_results,
            important_hits=important_hits,
            missing_items=missing_items,
            debug_logs=debug_logs,
        )

    def _apply_replacements(self, worksheets: List[Worksheet]) -> Tuple[ReplacementSummary, List[str]]:
        summary = ReplacementSummary()
        debug_logs: List[str] = []

        invalid_wb = load_workbook(self.config.invalid_part_db)
        invalid_ws = invalid_wb.active

        invalid_entries: Dict[str, Tuple[str, str, Optional[str], Optional[str]]] = {}
        for row in invalid_ws.iter_rows(min_row=2, values_only=True):
            invalid_no = str(row[0]).strip() if row[0] else ""
            invalid_desc = str(row[1]).strip() if row[1] else ""
            replacement_no = str(row[2]).strip() if row[2] else None
            replacement_desc = str(row[3]).strip() if row[3] else None
            if invalid_no:
                invalid_entries[normalize_part_no(invalid_no)] = (
                    invalid_no,
                    invalid_desc,
                    replacement_no,
                    replacement_desc,
                )

        for ws in worksheets:  # 遍历目标工作表，高亮并记录命中的失效料号
            part_col_idx = self._identify_part_column(ws)
            debug_logs.append(f"[{ws.title}] 识别料号列: {self._format_column_debug(part_col_idx)}")
            if part_col_idx is None:
                continue

            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if part_col_idx >= len(row):
                    continue
                cell_value = row[part_col_idx].value
                if not cell_value:
                    continue

                part_no = str(cell_value).strip()
                match = invalid_entries.get(normalize_part_no(part_no))
                if not match:
                    continue

                invalid_no, invalid_desc, replacement_no, replacement_desc = match

                part_cell = row[part_col_idx]
                summary.total_invalid_found += 1
                if self._row_already_replaced(row, part_col_idx, part_cell, replacement_no):
                    summary.total_invalid_previously_marked += 1
                    debug_logs.append(
                        f"[{ws.title}] 行{row_idx} 失效料号 {part_no} 已标记替换，跳过"
                    )
                    continue

                for cell in row:
                    cell.fill = BLACK_FILL

                if replacement_no:
                    replacement_col = ws.max_column + 1
                    ws.cell(row=row_idx, column=replacement_col).value = replacement_no
                    ws.cell(row=row_idx, column=replacement_col + 1).value = replacement_desc
                    summary.total_replaced += 1

                summary.records.append(
                    ReplacementRecord(
                        invalid_part_no=invalid_no,
                        invalid_desc=invalid_desc,
                        replacement_part_no=replacement_no,
                        replacement_desc=replacement_desc,
                        sheet_name=ws.title,
                        row_index=row_idx,
                    )
                )

                debug_logs.append(
                    f"[{ws.title}] 行{row_idx} 命中失效料号 {part_no} -> {replacement_no or '无替换'}"
                )

        return summary, debug_logs

    def _row_already_replaced(
        self,
        row: Tuple[Cell, ...],
        part_col_idx: int,
        part_cell: Cell,
        replacement_no: Optional[str],
    ) -> bool:
        if replacement_no and _row_contains_part(row, part_col_idx, replacement_no):
            return True
        if _is_black_fill(part_cell):
            if _row_has_non_black_value(row, part_col_idx):
                return True
            if not replacement_no:
                return True
        return False

    def _find_replacement_in_row(
        self, row: Tuple[Cell, ...], part_col_idx: int
    ) -> Optional[Tuple[str, str, Optional[str]]]:
        for idx, cell in enumerate(row):
            if idx == part_col_idx:
                continue
            value = cell.value
            if value in (None, ""):
                continue
            if isinstance(value, (int, float)):
                continue
            text = str(value).strip()
            if not text:
                continue
            if not _is_probable_part_number(text):
                continue
            if _is_black_fill(cell):
                continue
            normalized = normalize_part_no(text)
            desc_value: Optional[str] = None
            if idx + 1 < len(row):
                desc_cell = row[idx + 1]
                if desc_cell.value not in (None, ""):
                    desc_value = str(desc_cell.value).strip()
            return text, normalized, desc_value
        return None

    def _resolve_row_part(
        self, row: Tuple[Cell, ...], part_col_idx: int
    ) -> Optional[Tuple[str, str, Optional[str]]]:
        part_cell = row[part_col_idx]
        raw_value = part_cell.value
        text = str(raw_value).strip() if raw_value not in (None, "") else ""

        if text and not _is_black_fill(part_cell) and _is_probable_part_number(text):
            normalized = normalize_part_no(text)
            return normalized, text, None

        replacement = self._find_replacement_in_row(row, part_col_idx)
        if replacement:
            display_no, normalized, desc_value = replacement
            return normalized, display_no, desc_value

        if text and not _is_probable_part_number(text):
            return None

        if _is_black_fill(part_cell):
            return None

        if text:
            normalized = normalize_part_no(text)
            return normalized, text, None

        return None

    def _extract_part_quantities(
        self,
        worksheets: List[Worksheet],
    ) -> Tuple[Dict[str, float], Dict[str, str], Dict[str, str], List[str]]:
        part_quantities: Dict[str, float] = defaultdict(float)
        part_descriptions: Dict[str, str] = {}
        part_display: Dict[str, str] = {}
        debug_logs: List[str] = []

        skip_titles = {"执行统计", "剩余物料"}

        for ws in worksheets:  # 逐行累计第一个工作表中的库存数量与描述信息
            if ws.title in skip_titles:
                debug_logs.append(f"[{ws.title}] 已跳过汇总工作表")
                continue
            qty_col_idx = self._identify_quantity_column(ws)
            part_col_idx = self._identify_part_column(ws)
            desc_col_idx = self._identify_description_column(ws, part_col_idx)
            debug_logs.append(
                f"[{ws.title}] 数量列: {self._format_column_debug(qty_col_idx)}, 料号列: {self._format_column_debug(part_col_idx)}, 描述列: {self._format_column_debug(desc_col_idx)}"
            )

            if part_col_idx is None:
                continue

            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if part_col_idx >= len(row):
                    continue

                resolved = self._resolve_row_part(row, part_col_idx)
                if not resolved:
                    continue

                normalized_part, display_no, override_desc = resolved
                part_display.setdefault(normalized_part, display_no)


                normalized_part, display_no, override_desc = resolved
                part_display.setdefault(normalized_part, display_no)

                desc_value: Optional[str] = override_desc
                if not desc_value:
                    desc_cell = (
                        row[part_col_idx + 1]
                        if desc_col_idx is None and part_col_idx + 1 < len(row)
                        else None
                    )
                    if desc_col_idx is not None and desc_col_idx < len(row):
                        desc_cell = row[desc_col_idx]
                    if desc_cell and desc_cell.value:
                        desc_value = str(desc_cell.value).strip()
                if desc_value:
                    part_descriptions.setdefault(normalized_part, desc_value)

                quantity = 1.0
                if qty_col_idx is not None and qty_col_idx < len(row):
                    quantity_cell = row[qty_col_idx]
                    parsed_quantity = self._parse_quantity_value(quantity_cell.value)
                    if parsed_quantity is not None:
                        quantity = parsed_quantity
                    else:
                        quantity = 0.0
                        debug_logs.append(
                            f"[{ws.title}] 行{row_idx} 数量列值 {quantity_cell.value!r} 无法解析，按0处理"
                        )

                part_quantities[normalized_part] += quantity

        return part_quantities, part_descriptions, part_display, debug_logs

    def _identify_quantity_column(self, ws: Worksheet) -> Optional[int]:
        """Guess the quantity column by combining header keywords and numeric shape.

        The previous implementation required every non-empty cell to be parsable as a
        number, which fails when BOM sheets embed remarks such as "合计" or "-" inside
        the quantity column.  The revised heuristic keeps track of both successful and
        failed parses so that a mostly-numeric column is still selected.  The first two
        columns are ignored (常见为序号/料号)，并根据可解析为整数的单元格数量作为首要排序规则。
        """

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        header_candidates: List[int] = []
        if header_row:
            for idx, value in enumerate(header_row):
                if idx < 2:
                    continue
                if value in (None, ""):
                    continue
                lowered = str(value).strip().lower()
                if not lowered:
                    continue
                if any(keyword in lowered for keyword in ("数量", "數量", "qty", "quantity")):
                    header_candidates.append(idx)

        # (col_idx, integer_count, numeric_count, failure_count, total_count)
        numeric_scores: List[Tuple[int, int, int, int, int]] = []
        for col_idx in range(ws.max_column):
            if col_idx < 2:
                continue
            numeric_count = 0
            integer_count = 0
            failure_count = 0
            total_count = 0
            for cell in ws.iter_rows(
                min_row=2,
                min_col=col_idx + 1,
                max_col=col_idx + 1,
                values_only=True,
            ):
                value = cell[0]
                if value in (None, ""):
                    continue
                total_count += 1
                parsed = self._parse_quantity_value(value)
                if parsed is None:
                    failure_count += 1
                    continue
                numeric_count += 1
                if isclose(parsed, round(parsed), abs_tol=1e-6):
                    integer_count += 1
            if numeric_count:
                numeric_scores.append(
                    (col_idx, integer_count, numeric_count, failure_count, total_count)
                )

        def _select_best(
            scores: List[Tuple[int, int, int, int, int]]
        ) -> Optional[int]:
            if not scores:
                return None
            # 优先选择转换后整数数量最多的列，再比较可解析数量、失败数量以及列索引确保稳定。
            scores.sort(key=lambda item: (-item[1], -item[2], item[3], item[0]))
            return scores[0][0]

        if header_candidates:
            header_scores = [score for score in numeric_scores if score[0] in header_candidates]
            selected = _select_best(header_scores)
            if selected is not None:
                return selected
            # 如果表头标记了数量列但无有效数值，直接返回首个表头候选项，保持旧行为。
            # 如果表头标记了数量列但数据为空，直接返回表头候选项中的首个
            return header_candidates[0]

        return _select_best(numeric_scores)

    def _parse_quantity_value(self, value) -> Optional[float]:
        if value in (None, ""):
            return None
        if isinstance(value, bool):
            return float(value)
        if isinstance(value, (int, float)):
            try:
                number = float(value)
            except (TypeError, ValueError):
                return None
            if not (number == number):  # NaN check
                return None
            return number
        text = str(value).strip()
        if not text:
            return None
        normalized = text.replace(",", "")
        match = re.search(r"[-+]?\d+(?:\.\d+)?", normalized)
        if not match:
            return None
        try:
            return float(match.group())
        except ValueError:
            return None

    def _identify_part_column(self, ws: Worksheet) -> Optional[int]:
        candidate_scores: List[Tuple[int, int, int]] = []  # (idx, u_count, text_count)
        for col_idx in range(ws.max_column):
            u_count = 0
            text_count = 0
            for cell in ws.iter_rows(min_row=2, min_col=col_idx + 1, max_col=col_idx + 1, values_only=True):
                value = cell[0]
                if value is None:
                    continue
                text = str(value).strip()
                if not text:
                    continue
                text_count += 1
                if text.upper().startswith("U"):
                    u_count += 1
            if text_count:
                candidate_scores.append((col_idx, u_count, text_count))
        if not candidate_scores:
            return None
        candidate_scores.sort(key=lambda item: (-item[1], -item[2]))
        return candidate_scores[0][0]

    def _identify_description_column(self, ws: Worksheet, part_col_idx: Optional[int]) -> Optional[int]:
        for header_row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row)):
            for idx, cell in enumerate(header_row):
                value = cell.value if isinstance(cell, Cell) else cell
                if isinstance(value, str):
                    lowered = value.strip().lower()
                    if lowered and ("desc" in lowered or "描述" in lowered):
                        return idx
        if part_col_idx is not None and part_col_idx + 1 < ws.max_column:
            return part_col_idx + 1
        return None

    def _evaluate_binding_requirements(
        self,
        part_quantities: Dict[str, float],
        available_inventory: Dict[str, float],
        part_desc: Dict[str, str],
        part_display: Dict[str, str],
        binding_library: BindingLibrary,
    ) -> Tuple[List[BindingProjectResult], List[MissingItem], set[str], List[str]]:
        results: List[BindingProjectResult] = []
        missing_items: Dict[str, MissingItem] = {}
        used_parts: set[str] = set()
        debug_logs: List[str] = []

        for project in binding_library.iter_projects():
            index_key = normalize_part_no(project.index_part_no)
            project_qty = part_quantities.get(index_key, 0.0)
            if project_qty <= 0:
                continue

            available_index_qty = available_inventory.get(index_key, 0.0)
            if available_index_qty <= 0:
                continue

            consumption_qty = min(project_qty, available_index_qty)
            available_inventory[index_key] = max(available_index_qty - consumption_qty, 0.0)
            used_parts.add(index_key)

            debug_logs.append(
                f"[绑定]{project.project_desc}({project.index_part_no}) 主料需求: {project_qty} 可用: {available_index_qty}"
            )

            group_results: List[RequirementGroupResult] = []
            for group in project.required_groups:
                result = self._evaluate_group(
                    group,
                    consumption_qty,
                    available_inventory,
                    part_quantities,
                    part_display,
                )
                group_results.append(result)

                if result.missing_qty > 0:
                    for part_no in result.missing_choices:
                        part_key = normalize_part_no(part_no)
                        description = part_desc.get(part_key) or self._lookup_choice_desc(group, part_no)
                        display_no = part_display.get(part_key, part_no)
                        item = missing_items.setdefault(
                            part_key,
                            MissingItem(part_no=display_no, desc=description, missing_qty=0.0),
                        )
                        if not item.desc and description:
                            item.desc = description
                        item.missing_qty += result.missing_qty

                for matched_part_no in result.matched_details.keys():
                    used_parts.add(normalize_part_no(matched_part_no))

            results.append(
                BindingProjectResult(
                    project_desc=project.project_desc,
                    index_part_no=project.index_part_no,
                    matched_quantity=consumption_qty,
                    requirement_results=group_results,
                )
            )

        return results, list(missing_items.values()), used_parts, debug_logs

    def _evaluate_group(
        self,
        group: BindingGroup,
        project_qty: float,
        available_inventory: Dict[str, float],
        reference_quantities: Dict[str, float],
        part_display: Dict[str, str],
    ) -> RequirementGroupResult:
        base_requirement = group.number if group.number not in (None, "") else 1.0
        try:
            base_requirement = float(base_requirement)
        except (TypeError, ValueError):
            base_requirement = 1.0

        required_qty = project_qty * base_requirement
        available_qty = 0.0
        fulfilled_qty = 0.0
        matched_details: Dict[str, float] = {}
        applicable_choices: List[Tuple[int, BindingChoice, str, float]] = []
        fallback_choices: List[str] = []
        first_applicable_part: Optional[str] = None
        requirement_enabled = False

        for idx, choice in enumerate(group.choices):
            if not choice.part_no:
                continue
            if not self._choice_condition_met(choice, reference_quantities):
                continue

            requirement_enabled = True
            fallback_choices.append(choice.part_no)

            choice_key = normalize_part_no(choice.part_no)
            total_stock = reference_quantities.get(choice_key, 0.0)
            if total_stock > 0:
                available_qty += total_stock
            stock = max(available_inventory.get(choice_key, 0.0), 0.0)
            applicable_choices.append((idx, choice, choice_key, stock))
            if first_applicable_part is None:
                first_applicable_part = choice.part_no

        if not requirement_enabled:
            return RequirementGroupResult(
                group_name=group.group_name,
                required_qty=0.0,
                available_qty=0.0,
                missing_qty=0.0,
                missing_choices=[],
                matched_details={},
            )

        applicable_choices.sort(key=lambda item: (-item[3], item[0]))

        for _idx, choice, choice_key, stock in applicable_choices:
            remaining_need = max(required_qty - fulfilled_qty, 0.0)
            if remaining_need <= 0:
                break

            current_stock = max(available_inventory.get(choice_key, 0.0), 0.0)
            if current_stock <= 0:
                continue

            take_amount = min(current_stock, remaining_need)
            if take_amount <= 0:
                continue

            display_no = part_display.get(choice_key, choice.part_no)
            matched_details[display_no] = matched_details.get(display_no, 0.0) + take_amount
            fulfilled_qty += take_amount
            available_inventory[choice_key] = max(current_stock - take_amount, 0.0)

        missing_qty = max(required_qty - fulfilled_qty, 0.0)
        missing_choices: List[str] = []
        if missing_qty > 0:
            if first_applicable_part:
                missing_choices = [first_applicable_part]
            elif fallback_choices:
                missing_choices = [fallback_choices[0]]
            if not missing_choices:
                missing_choices = [choice.part_no for choice in group.choices if choice.part_no]

        return RequirementGroupResult(
            group_name=group.group_name,
            required_qty=required_qty,
            available_qty=available_qty,
            missing_qty=missing_qty,
            missing_choices=missing_choices,
            matched_details=matched_details,
        )

    def _lookup_choice_desc(self, group: BindingGroup, part_no: str) -> str:
        for choice in group.choices:
            if choice.part_no == part_no and choice.desc:
                return choice.desc
        return ""

    def _choice_condition_met(self, choice, part_quantities: Dict[str, float]) -> bool:
        mode = (choice.condition_mode or "").upper()
        if not mode:
            return True

        condition_keys = [
            normalize_part_no(part_no)
            for part_no in choice.condition_part_nos
            if part_no not in (None, "")
        ]
        if not condition_keys:
            return False

        def _has_part(part_no: str) -> bool:
            quantity = part_quantities.get(part_no, 0.0)
            return bool(quantity and quantity > 0)

        if mode == "ALL":
            return all(_has_part(part_no) for part_no in condition_keys)
        if mode == "ANY":
            return any(_has_part(part_no) for part_no in condition_keys)
        if mode == "NOTANY":
            return not any(_has_part(part_no) for part_no in condition_keys)
        return True

    def _scan_important_materials(
        self,
        part_quantities: Dict[str, float],
        part_display: Dict[str, str],
        part_descriptions: Dict[str, str],
    ) -> Tuple[List[ImportantMaterialHit], set[str], List[str]]:
        important_path = self.config.important_materials
        hits: List[ImportantMaterialHit] = []
        matched_parts: set[str] = set()
        debug_logs: List[str] = []

        if not important_path.exists():
            return hits, matched_parts, debug_logs

        keywords = [
            line.strip()
            for line in important_path.read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]
        part_variant_cache: Dict[str, set[str]] = {}

        for keyword in keywords:
            normalized_keyword = normalize_text(keyword)
            keyword_variants = normalized_variants(keyword)
            if normalized_keyword:
                keyword_variants.add(normalized_keyword)

            total_qty = 0.0
            matched_detail: Dict[str, float] = {}

            for part_no, qty in part_quantities.items():
                variants = part_variant_cache.get(part_no)
                if variants is None:
                    variants = self._collect_part_variants(
                        part_display.get(part_no, part_no),
                        part_no,
                        part_descriptions.get(part_no, ""),
                    )
                    part_variant_cache[part_no] = variants

                if not self._variants_match(keyword_variants, variants):
                    continue

                display_no = part_display.get(part_no, part_no)
                total_qty += qty
                matched_detail[display_no] = matched_detail.get(display_no, 0.0) + qty
                matched_parts.add(part_no)

            if total_qty:
                hits.append(
                    ImportantMaterialHit(
                        keyword=keyword,
                        converted_keyword=normalized_keyword,
                        total_quantity=total_qty,
                        matched_parts=matched_detail,
                    )
                )
                debug_logs.append(
                    f"[重要物料] {keyword} 命中 {len(matched_detail)} 项，数量 {total_qty}"
                )
            else:
                debug_logs.append(f"[重要物料] {keyword} 未命中")

        return hits, matched_parts, debug_logs

    def _write_result_sheets(
        self,
        wb,
        replacement_summary: ReplacementSummary,
        binding_results: List[BindingProjectResult],
        important_hits: List[ImportantMaterialHit],
        missing_items: List[MissingItem],
        part_quantities: Dict[str, float],
        part_descriptions: Dict[str, str],
        part_display: Dict[str, str],
        remainder_keys: set[str],
        debug_logs: List[str],
    ) -> None:
        if "执行统计" in wb.sheetnames:
            del wb["执行统计"]
        if "剩余物料" in wb.sheetnames:
            del wb["剩余物料"]

        summary_ws = wb.create_sheet("执行统计")
        summary_ws.append(["失效料号数量", replacement_summary.total_invalid_found])
        summary_ws.append([
            "已标记失效料号数量",
            replacement_summary.total_invalid_previously_marked,
        ])
        summary_ws.append(["已替换数量", replacement_summary.total_replaced])
        summary_ws.append(["绑定项目数量", len(binding_results)])
        binding_group_count = sum(len(result.requirement_results) for result in binding_results)
        summary_ws.append(["绑定分组数量", binding_group_count])
        summary_ws.append(["重要物料数量", len(important_hits)])

        summary_ws.append([])
        summary_ws.append(["失效料号明细"])
        summary_ws.append(["工作表", "行号", "失效料号", "失效描述", "替换料号", "替换描述"])
        for record in replacement_summary.records:
            summary_ws.append(
                [
                    record.sheet_name,
                    record.row_index,
                    record.invalid_part_no,
                    record.invalid_desc,
                    record.replacement_part_no or "",
                    record.replacement_desc or "",
                ]
            )

        summary_ws.append([])
        summary_ws.append(["绑定料号统计"])
        summary_ws.append([
            "项目描述",
            "索引料号",
            "主料数量",
            "需求分组",
            "需求数量",
            "可用数量",
            "缺少数量",
            "缺少料号",
            "满足料号",
        ])
        for result in binding_results:
            for group_result in result.requirement_results:
                matched_parts_text = ",".join(
                    f"{part}:{format_quantity_text(qty)}"
                    for part, qty in group_result.matched_details.items()
                )
                summary_ws.append(
                    [
                        result.project_desc,
                        result.index_part_no,
                        format_quantity_cell(result.matched_quantity),
                        group_result.group_name,
                        format_quantity_cell(group_result.required_qty),
                        format_quantity_cell(group_result.available_qty),
                        format_quantity_cell(group_result.missing_qty),
                        ",".join(group_result.missing_choices),
                        matched_parts_text,
                    ]
                )

        summary_ws.append([])
        summary_ws.append(["缺失物料"])
        summary_ws.append(["料号", "描述", "缺少数量"])
        for item in missing_items:
            summary_ws.append(
                [item.part_no, item.desc, format_quantity_cell(item.missing_qty)]
            )

        summary_ws.append([])
        summary_ws.append(["重要物料统计"])
        summary_ws.append(["关键字", "标准关键字", "数量", "命中料号"])
        for hit in important_hits:
            matched_text = ",".join(
                f"{part}:{format_quantity_text(qty)}"
                for part, qty in hit.matched_parts.items()
            )
            summary_ws.append(
                [
                    hit.keyword,
                    hit.converted_keyword,
                    format_quantity_cell(hit.total_quantity),
                    matched_text,
                ]
            )

        summary_ws.append([])
        summary_ws.append(["调试信息"])
        for line in debug_logs:
            summary_ws.append([line])

        remainder_ws = wb.create_sheet("剩余物料")
        remainder_ws.append(["料号", "描述", "数量"])
        for key in sorted(remainder_keys, key=lambda k: part_display.get(k, k)):
            remainder_ws.append(
                [
                    part_display.get(key, key),
                    part_descriptions.get(key, ""),
                    format_quantity_cell(part_quantities.get(key, 0.0)),
                ]
            )

    def _collect_part_variants(self, display_no: str, part_no: str, description: str) -> set[str]:
        variants: set[str] = set()
        variants.update(normalized_variants(display_no))
        variants.update(normalized_variants(part_no))
        if description:
            variants.update(normalized_variants(description))
        return variants

    def _variants_match(self, keyword_variants: set[str], value_variants: set[str]) -> bool:
        if not keyword_variants or not value_variants:
            return False
        for keyword_variant in keyword_variants:
            if not keyword_variant:
                continue
            for value_variant in value_variants:
                if not value_variant:
                    continue
                if keyword_variant in value_variant or value_variant in keyword_variant:
                    return True
        return False

    def _format_column_debug(self, col_idx: Optional[int]) -> str:
        if col_idx is None:
            return "未识别"
        return f"第{col_idx + 1}列"
