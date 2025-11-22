from __future__ import annotations

from dataclasses import dataclass, field
from math import isclose
import csv
import re
from pathlib import Path
from typing import Dict

from openpyxl import Workbook, load_workbook

from .excel_processor import format_quantity_text, normalize_part_no
from .text_utils import normalize_text, normalized_variants


@dataclass
class SystemPartRecord:
    part_no: str
    description: str
    unit: str
    applicant: str
    inventory: float | None
    _categories: tuple[str, ...] = field(init=False, repr=False)
    _search_fields: tuple[str, ...] = field(init=False, repr=False)

    def __post_init__(self) -> None:
        categories: list[str] = [_categorize_part_no(self.part_no)]
        parts = [segment.strip() for segment in self.description.split(";") if segment.strip()]
        categories.extend(parts)
        self._categories = tuple(categories) if categories else ("未分类",)
        self._search_fields = _prepare_search_fields(self.part_no, self.description, self.applicant)

    @property
    def categories(self) -> tuple[str, ...]:
        return self._categories

    @property
    def search_fields(self) -> tuple[str, ...]:
        return self._search_fields

    @property
    def inventory_display(self) -> str:
        if self.inventory is None:
            return ""
        if isclose(self.inventory, round(self.inventory), abs_tol=1e-6):
            return str(int(round(self.inventory)))
        return f"{round(self.inventory, 4):g}"


class SystemPartRepository:
    def __init__(self, path: Path) -> None:
        self.path = path
        self.records: list[SystemPartRecord] = []
        self._index: dict[str, SystemPartRecord] = {}
        self.load()

    def load(self) -> None:
        if not self.path.exists():
            raise FileNotFoundError(f"系统料号文件不存在：{self.path}")
        source = self._prefer_fast_path()
        records = _parse_system_parts(source)
        index: dict[str, SystemPartRecord] = {}
        normalized_records: list[SystemPartRecord] = []
        for record in records:
            normalized = normalize_part_no(record.part_no)
            if not normalized:
                continue
            normalized_records.append(record)
            if normalized not in index:
                index[normalized] = record

        self.records = normalized_records
        self._index = index

    def _prefer_fast_path(self) -> Path:
        suffix = self.path.suffix.lower()
        if suffix in {".xlsx", ".xlsm"}:
            cached = self.path.with_suffix(".tsv")
            try:
                if (not cached.exists()) or cached.stat().st_mtime < self.path.stat().st_mtime:
                    self._convert_excel_to_tsv(self.path, cached)
                return cached
            except Exception:
                return self.path
        return self.path

    def _convert_excel_to_tsv(self, source: Path, destination: Path) -> None:
        workbook = load_workbook(source, data_only=True, read_only=True)
        sheet = workbook.active
        destination.parent.mkdir(parents=True, exist_ok=True)
        with destination.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle, delimiter="\t")
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(["" if cell is None else cell for cell in row])
        workbook.close()

    def find(self, part_no: str) -> SystemPartRecord | None:
        normalized = normalize_part_no(part_no)
        if not normalized:
            return None
        return self._index.get(normalized)

    def build_hierarchy(self, query: str | None = None) -> Dict[str, Dict]:
        keywords = _prepare_keywords(query)
        root: Dict[str, Dict] = {"children": {}, "parts": []}

        for record in self.records:
            if keywords and not _matches_query(record, keywords):
                continue
            node = root
            for category in record.categories:
                node = node["children"].setdefault(category, {"children": {}, "parts": []})
            node.setdefault("parts", []).append(record)

        return root

    def search(self, query: str) -> list[SystemPartRecord]:
        keywords = _prepare_keywords(query)
        if not keywords:
            return list(self.records)
        return [record for record in self.records if _matches_query(record, keywords)]


def generate_system_part_exports(
    source_path: Path,
    invalid_part_path: Path,
    blocked_applicant_path: Path,
) -> tuple[Path, Path]:
    if not source_path.exists():
        raise FileNotFoundError(f"找不到系统料号原始文件：{source_path}")

    invalid_part_numbers = _load_invalid_part_numbers(invalid_part_path)
    blocked_applicants = _load_blocked_applicants(blocked_applicant_path)

    records = _parse_system_parts(source_path)

    filtered_records: list[SystemPartRecord] = []
    seen_parts: set[str] = set()
    for record in records:
        normalized_part = normalize_part_no(record.part_no)
        if not normalized_part.startswith("UC3"):
            continue
        if normalized_part in invalid_part_numbers:
            continue
        if _should_block(record, blocked_applicants):
            continue
        if normalized_part in seen_parts:
            continue
        seen_parts.add(normalized_part)
        filtered_records.append(record)

    destination = source_path.with_suffix(".xlsx")
    if destination == source_path:
        destination = source_path.with_name(source_path.name + ".xlsx")

    destination.parent.mkdir(parents=True, exist_ok=True)
    _write_excel(filtered_records, destination)
    fast_path = destination.with_suffix(".tsv")
    _write_tsv(filtered_records, fast_path)
    return destination, fast_path


def generate_system_part_excel(
    source_path: Path,
    invalid_part_path: Path,
    blocked_applicant_path: Path,
) -> Path:
    excel_path, _ = generate_system_part_exports(
        source_path, invalid_part_path, blocked_applicant_path
    )
    return excel_path


def _write_excel(records: list[SystemPartRecord], destination: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "系统料号"
    sheet.append(["料号", "描述", "单位", "申请人", "库存"])

    for record in records:
        sheet.append(
            [
                record.part_no,
                record.description,
                record.unit,
                record.applicant,
                _format_inventory_cell(record.inventory),
            ]
        )

    workbook.save(destination)
    workbook.close()


def _write_tsv(records: list[SystemPartRecord], destination: Path) -> None:
    with destination.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t")
        writer.writerow(["料号", "描述", "单位", "申请人", "库存"])
        for record in records:
            writer.writerow(
                [
                    record.part_no,
                    record.description,
                    record.unit,
                    record.applicant,
                    _format_inventory_cell(record.inventory),
                ]
            )


def _parse_system_parts(path: Path) -> list[SystemPartRecord]:
    suffix = path.suffix.lower()
    if suffix in {".tsv", ".txt"}:
        return _parse_tsv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _parse_excel(path)
    raise ValueError(f"不支持的系统料号文件格式：{path.suffix}")


def _parse_tsv(path: Path) -> list[SystemPartRecord]:
    records: list[SystemPartRecord] = []
    with path.open("r", encoding="utf-8") as handle:
        reader = csv.reader(handle, delimiter="\t")
        for row in reader:
            if not row:
                continue
            if _safe_str(row[0]).lower() in {"料号", "part_no"}:
                continue
            if len(row) >= 11:
                part_no = _safe_str(row[1])
                description = _safe_str(row[3])
                unit = _safe_str(row[6])
                applicant = _clean_applicant_text(_safe_str(row[9]))
                inventory_raw = row[10]
            elif len(row) >= 5:
                part_no = _safe_str(row[0])
                description = _safe_str(row[1])
                unit = _safe_str(row[2])
                applicant = _clean_applicant_text(_safe_str(row[3]))
                inventory_raw = row[4]
            else:
                continue
            if not part_no:
                continue
            inventory = _convert_inventory(inventory_raw)
            records.append(SystemPartRecord(part_no, description, unit, applicant, inventory))
    return records


def _parse_excel(path: Path) -> list[SystemPartRecord]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    records: list[SystemPartRecord] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        part_no = _safe_str(row[0])
        description = _safe_str(row[1])
        if not part_no:
            continue
        unit = _safe_str(row[2])
        applicant = _clean_applicant_text(_safe_str(row[3]))
        inventory = _convert_inventory(row[4] if len(row) > 4 else None)
        records.append(SystemPartRecord(part_no, description, unit, applicant, inventory))
    workbook.close()
    return records


def _categorize_part_no(part_no: str) -> str:
    normalized = normalize_part_no(part_no)
    if normalized.startswith("UC1"):
        return "加工件"
    if normalized.startswith("UC2"):
        return "机构外购件"
    if normalized.startswith("UC3"):
        return "电控外购件"
    if normalized.startswith("UA"):
        return "成品"
    if normalized.startswith("UB"):
        return "半成品"
    return "未分类"


def _convert_inventory(value) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        cleaned = str(value).strip()
        try:
            return float(cleaned)
        except ValueError:
            return None


def _format_inventory_cell(value: float | None):
    if value is None:
        return ""
    formatted = format_quantity_text(value)
    try:
        return float(formatted)
    except (TypeError, ValueError):
        return formatted


def _load_invalid_part_numbers(path: Path) -> set[str]:
    if not path.exists():
        return set()
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    invalid_numbers: set[str] = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        invalid_no = _safe_str(row[0])
        if invalid_no:
            invalid_numbers.add(normalize_part_no(invalid_no))
    workbook.close()
    return invalid_numbers


def _load_blocked_applicants(path: Path) -> "BlockedApplicantMatcher":
    matcher = BlockedApplicantMatcher()
    if not path.exists():
        return matcher
    content = path.read_text(encoding="utf-8")
    tokens = re.split(r"[\s,;，；]+", content)
    for token in tokens:
        matcher.add(token)
    return matcher


def _should_block(record: SystemPartRecord, blocked: "BlockedApplicantMatcher") -> bool:
    if not blocked:
        return False
    return blocked.matches(record.applicant, record.description)


def _matches_query(record: SystemPartRecord, keywords: list[set[str]]) -> bool:
    if not keywords:
        return True
    for keyword_variants in keywords:
        if not keyword_variants:
            continue
        matched = False
        for text in record.search_fields:
            for variant in keyword_variants:
                if variant and variant in text:
                    matched = True
                    break
            if matched:
                break
        if not matched:
            return False
    return True


def _prepare_search_fields(part_no: str, description: str, applicant: str) -> tuple[str, ...]:
    fields: list[str] = []

    def add(value: str) -> None:
        if value and value not in fields:
            fields.append(value)

    for value in (part_no, description, applicant):
        base = (value or "").strip().lower()
        add(base)
        normalized = normalize_text(value)
        add(normalized)

    normalized_part = normalize_part_no(part_no)
    if normalized_part:
        add(normalized_part.lower())

    return tuple(fields)


def _safe_str(value) -> str:
    return str(value).strip() if value not in (None, "") else ""


def _clean_applicant_text(value: str) -> str:
    return value.strip().strip(",，;；")


class BlockedApplicantMatcher:
    def __init__(self) -> None:
        self._variant_lengths: dict[str, set[int]] = {}
        self._description_tokens: set[str] = set()

    def add(self, value: str) -> None:
        if value is None:
            return
        cleaned = value.strip()
        if not cleaned:
            return
        if cleaned[0] in {"-", "－", "–", "—"}:
            target = cleaned[1:].strip()
            for variant in normalized_variants(target):
                if variant:
                    self._description_tokens.add(variant)
            return
        length = len(cleaned)
        for variant in normalized_variants(cleaned):
            if not variant:
                continue
            self._variant_lengths.setdefault(variant, set()).add(length)

    def matches(self, applicant: str, description: str | None = None) -> bool:
        if description and self._description_tokens:
            if self._matches_description(description):
                return True
        if not self._variant_lengths or not applicant:
            return False
        for token in _split_applicant_tokens(applicant):
            token_length = len(token)
            for variant in normalized_variants(token):
                lengths = self._variant_lengths.get(variant)
                if not lengths:
                    continue
                if token_length == 2:
                    if 2 in lengths:
                        return True
                else:
                    return True
        return False

    def _matches_description(self, description: str) -> bool:
        variants = normalized_variants(description)
        base = description.strip().lower()
        if base:
            variants.add(base)
        for text in variants:
            for token in self._description_tokens:
                if token and token in text:
                    return True
        return False


def _split_applicant_tokens(value: str) -> list[str]:
    raw_tokens = re.split(r"[\s,;，；/、\|]+", value)
    tokens: list[str] = []
    for token in raw_tokens:
        cleaned = token.strip().strip("()（）[]【】'\"")
        if cleaned:
            tokens.append(cleaned)
    return tokens


def _prepare_keywords(query: str | None) -> list[set[str]]:
    if not query:
        return []
    tokens = [segment.strip() for segment in re.split(r"[\s,;，；]+", query)]
    keyword_sets: list[set[str]] = []
    for token in tokens:
        if not token:
            continue
        variants = set(normalized_variants(token))
        normalized = normalize_text(token)
        if normalized:
            variants.add(normalized)
        base = token.strip().lower()
        if base:
            variants.add(base)
        keyword_sets.append({variant for variant in variants if variant})
    return keyword_sets

