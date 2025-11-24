from __future__ import annotations

import csv
import json
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Optional
from urllib.parse import urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from .excel_processor import normalize_part_no
from .part_assets import PartAsset, PartAssetStore


USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0 Safari/537.36"
)


@dataclass
class CrawlStatus:
    part_no: str
    status: str = "pending"  # pending | done | failed
    message: str = ""

    def to_dict(self) -> Dict:
        return {
            "part_no": self.part_no,
            "status": self.status,
            "message": self.message,
        }

    @classmethod
    def from_dict(cls, data: Dict) -> "CrawlStatus":
        return cls(
            part_no=data.get("part_no", ""),
            status=data.get("status", "pending"),
            message=data.get("message", ""),
        )


class AssetCrawler:
    def __init__(
        self,
        asset_root: Path,
        progress_path: Optional[Path] = None,
        delay_seconds: float = 1.0,
        description_lookup: Optional[Callable[[str], str]] = None,
        ua_lookup_dir: Optional[Path] = None,
    ) -> None:
        self.store = PartAssetStore(asset_root)
        self.progress_path = progress_path or (asset_root / "crawl_progress.json")
        self.delay_seconds = delay_seconds
        self._description_lookup = description_lookup
        self._ua_lookup_dir = ua_lookup_dir if ua_lookup_dir and ua_lookup_dir.exists() else None
        self._ua_sources: list[Path] = []
        self._tasks: Dict[str, CrawlStatus] = {}
        if self._ua_lookup_dir:
            self._ua_sources = self._collect_ua_sources(self._ua_lookup_dir)
        self._load_progress()

    def _load_progress(self) -> None:
        if not self.progress_path.exists():
            return
        try:
            raw = json.loads(self.progress_path.read_text(encoding="utf-8"))
            for item in raw:
                status = CrawlStatus.from_dict(item)
                if status.part_no:
                    self._tasks[status.part_no] = status
        except json.JSONDecodeError:
            # 如果进度文件损坏，则忽略并重新开始
            self._tasks = {}

    def _save_progress(self) -> None:
        payload = [task.to_dict() for task in self._tasks.values()]
        self.progress_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
        )

    def add_tasks(self, part_numbers: Iterable[str]) -> None:
        changed = False
        for part in part_numbers:
            normalized = normalize_part_no(part)
            if not normalized:
                continue
            existing = self._tasks.get(normalized)
            if existing:
                if existing.status == "done":
                    existing.status = "pending"
                    existing.message = ""
                    changed = True
                continue
            self._tasks[normalized] = CrawlStatus(part_no=normalized)
            changed = True
        if changed:
            self._save_progress()

    def remove_tasks(self, part_numbers: Iterable[str]) -> None:
        removed = False
        for part in part_numbers:
            normalized = normalize_part_no(part)
            if normalized and normalized in self._tasks:
                del self._tasks[normalized]
                removed = True
        if removed:
            self._save_progress()

    def clear(self) -> None:
        if not self._tasks:
            return
        self._tasks = {}
        self._save_progress()

    def pending(self) -> List[str]:
        return [p for p, task in self._tasks.items() if task.status != "done"]

    def run(self, limit: Optional[int] = None, should_cancel=None) -> bool:
        processed = 0
        cancelled = False
        for part_no in list(self.pending()):
            if should_cancel and should_cancel():
                cancelled = True
                break
            if limit is not None and processed >= limit:
                break
            status = self._tasks[part_no]
            try:
                message = self._process_part(part_no)
                status.status = "done"
                status.message = message
            except Exception as exc:  # noqa: BLE001
                status.status = "failed"
                status.message = str(exc)
            self._tasks[part_no] = status
            self._save_progress()
            processed += 1
            if self.delay_seconds:
                time.sleep(self.delay_seconds)
        return cancelled

    def statuses(self) -> List[CrawlStatus]:
        return sorted(self._tasks.values(), key=lambda item: item.part_no)

    def summary(self) -> tuple[int, int]:
        total = len(self._tasks)
        done = len([t for t in self._tasks.values() if t.status == "done"])
        return done, total

    def _process_part(self, part_no: str) -> str:
        normalized = normalize_part_no(part_no) or part_no
        if normalized.startswith("UB"):
            return "UB 料号无需自动生成，已跳过"

        updates: list[str] = []
        existing_asset = self.store.get(part_no)
        should_overwrite_uc = normalized.startswith("UC") and existing_asset is not None

        description = self._lookup_description(part_no)
        brand, model = _extract_brand_model(description)
        search_terms = _build_search_terms(part_no, description, brand, model)

        if normalized.startswith("UA"):
            updates.extend(self._update_from_ua_sources(normalized))
        else:
            primary_keyword = " ".join(filter(None, (brand, model))) or part_no
            official = self._search_official_site(primary_keyword)
            if official:
                asset = existing_asset or PartAsset(part_no=part_no)
                if should_overwrite_uc:
                    updated_links = [official]
                else:
                    updated_links = list(asset.remote_links)
                    if official not in updated_links:
                        updated_links.append(official)
                if updated_links != asset.remote_links:
                    self.store.set_remote_links(part_no, updated_links)
                    updates.append("官网链接")

        if not (existing_asset.images if existing_asset else []):
            for keyword in search_terms:
                image_path = self.store.download_first_image_from_search(part_no, keyword)
                if image_path:
                    updates.append("图片")
                    break

        elif should_overwrite_uc:
            for keyword in search_terms:
                image_path = self.store.download_first_image_from_search(part_no, keyword)
                if image_path:
                    asset = self.store.get(part_no)
                    if asset:
                        asset.images = [image_path]
                        self.store.upsert(asset)
                    updates.append("图片")
                    break

        if not updates:
            if existing_asset:
                return "已存在资源，未更新"
            return "未找到可保存的资源"
        return f"已更新 {', '.join(updates)}"

    def _lookup_description(self, part_no: str) -> str:
        if not self._description_lookup:
            return ""
        try:
            return self._description_lookup(part_no) or ""
        except Exception:
            return ""

    def _search_official_site(self, keyword: str) -> Optional[str]:
        response = requests.get(
            "https://www.bing.com/search",
            params={"q": f"{keyword} 官网", "setlang": "zh-cn"},
            headers={"User-Agent": USER_AGENT},
            timeout=15,
        )
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")
        normalized = (normalize_part_no(keyword) or keyword).lower()
        fallback: Optional[str] = None
        for link in soup.select("li.b_algo h2 a, ol#b_results h2 a"):
            href = link.get("href")
            if not href or not self._is_http_url(href):
                continue
            if normalized in href.lower():
                return href
            if fallback is None:
                fallback = href
        return fallback

    def _is_http_url(self, url: str) -> bool:
        parsed = urlparse(url)
        return parsed.scheme in {"http", "https"} and bool(parsed.netloc)

    def _update_from_ua_sources(self, part_no: str) -> list[str]:
        if not self._ua_sources:
            return []

        found_local: list[str] = []
        for path in self._ua_sources:
            suffix = path.suffix.lower()
            if suffix in {".xlsx", ".xlsm", ".xls"}:
                local = self._search_in_excel(path, part_no)
            elif suffix in {".csv", ".txt"}:
                local = self._search_in_csv(path, part_no)
            found_local.extend(local)

        updates: list[str] = []
        found_local = list(dict.fromkeys(found_local))

        if found_local:
            asset = self.store.get(part_no) or PartAsset(part_no=part_no)
            existing_local = set(asset.local_paths)
            merged_local = list(existing_local)
            for path in found_local:
                if path not in existing_local:
                    merged_local.append(path)
            if merged_local != list(existing_local):
                self.store.set_local_paths(part_no, merged_local)
                updates.append("UA档案")

        return updates

    def _search_in_excel(self, path: Path, part_no: str) -> list[str]:
        local: list[str] = []
        normalized = normalize_part_no(part_no) or part_no
        try:
            workbook = load_workbook(path, data_only=True, read_only=True)
        except Exception:
            return local

        try:
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    if not row:
                        continue
                    if any(self._cell_contains_part(cell, normalized) for cell in row):
                        local.extend(self._extract_local_paths_from_row(row))
        finally:
            workbook.close()

        return local

    def _search_in_csv(self, path: Path, part_no: str) -> list[str]:
        local: list[str] = []
        normalized = normalize_part_no(part_no) or part_no
        try:
            with path.open("r", encoding="utf-8", errors="ignore") as handle:
                reader = csv.reader(handle)
                for row in reader:
                    if not row:
                        continue
                    if any(self._cell_contains_part(cell, normalized) for cell in row):
                        local.extend(self._extract_local_paths_from_row(row))
        except Exception:
            return local

        return local

    def _collect_ua_sources(self, root: Path) -> list[Path]:
        sources: list[Path] = []
        supported = {".xlsx", ".xlsm", ".xls", ".csv", ".txt"}
        for path in root.rglob("*"):
            if not path.is_file():
                continue
            if path.suffix.lower() not in supported:
                continue
            if path.name.startswith("~$"):
                continue
            sources.append(path)
        sources.sort()
        return sources

    def _cell_contains_part(self, value, normalized_part: str) -> bool:
        if value is None:
            return False
        text = str(value).strip()
        normalized_value = normalize_part_no(text) or text
        lower_value = normalized_value.lower()
        lower_part = normalized_part.lower()
        return lower_part == lower_value or lower_part in lower_value

    def _extract_local_paths_from_row(self, values: Iterable) -> list[str]:
        paths: list[str] = []
        for value in values:
            if value is None:
                continue
            text = str(value).strip()
            if not text or self._is_http_url(text):
                continue
            if "\\" in text or "/" in text:
                cleaned = text
                if cleaned.startswith("\\") and not cleaned.startswith("\\\\"):
                    cleaned = "\\" + cleaned
                cleaned = cleaned.strip()
                if cleaned and cleaned not in paths:
                    paths.append(cleaned)
        return paths

    def _extract_http_links(self, values: Iterable) -> list[str]:
        links: list[str] = []
        for value in values:
            if value is None:
                continue
            for match in re.findall(r"https?://[^\s]+", str(value)):
                cleaned = match.strip().rstrip(",.;)\"]")
                if cleaned and cleaned not in links:
                    links.append(cleaned)
        return links


def _extract_brand_model(description: str) -> tuple[str | None, str | None]:
    brand = _extract_labeled_value(description, ("品牌", "牌子", "厂家", "厂商"))
    model = _extract_labeled_value(description, ("型号", "规格型号", "机型"))

    tokens = [token for token in re.split(r"[\s,;，；/、]+", description or "") if token]
    if not brand and tokens:
        brand = tokens[0]
    if not model and len(tokens) > 1:
        model = tokens[1]

    return brand, model


def _extract_labeled_value(description: str, labels: tuple[str, ...]) -> str | None:
    for label in labels:
        match = re.search(rf"{label}\s*[:：]?\s*([^,;；，/\s]+)", description or "")
        if match:
            value = match.group(1).strip()
            if value:
                return value
    return None


def _build_search_terms(
    part_no: str, description: str, brand: str | None, model: str | None
) -> list[str]:
    terms: list[str] = []
    base_pairs = [" ".join(filter(None, (brand, model))), model, description]
    for phrase in base_pairs:
        if not phrase:
            continue
        for suffix in (" 产品 图片", " 图片", ""):
            keyword = f"{phrase}{suffix}".strip()
            if keyword and keyword not in terms:
                terms.append(keyword)

    for keyword in (f"{part_no} 产品 图片", part_no):
        if keyword not in terms:
            terms.append(keyword)

    return terms


__all__ = ["AssetCrawler", "CrawlStatus"]
