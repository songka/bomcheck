from __future__ import annotations

import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import Workbook, load_workbook


@dataclass
class BindingChoice:
    part_no: str
    desc: str
    condition_mode: Optional[str] = None
    condition_part_nos: List[str] = field(default_factory=list)
    number: Optional[float] = None

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "BindingChoice":
        return cls(
            part_no=data.get("partNo", ""),
            desc=data.get("desc", ""),
            condition_mode=data.get("conditionMode"),
            condition_part_nos=list(data.get("conditionPartNos", []) or []),
            number=_parse_number(data.get("number")),
        )

    def to_dict(self) -> Dict[str, Any]:
        data: Dict[str, Any] = {
            "partNo": self.part_no,
            "desc": self.desc,
        }
        if self.condition_mode:
            data["conditionMode"] = self.condition_mode
        if self.condition_part_nos:
            data["conditionPartNos"] = self.condition_part_nos
        if self.number is not None:
            data["number"] = self.number
        return data


@dataclass
class BindingGroup:
    group_name: str
    number: float = 1.0
    choices: List[BindingChoice] = field(default_factory=list)

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "BindingGroup":
        return cls(
            group_name=data.get("groupName", ""),
            number=_parse_number(data.get("number", 1)) or 1.0,
            choices=[BindingChoice.from_dict(item) for item in data.get("choices", [])],
        )

    def to_dict(self) -> Dict[str, Any]:
        data: Dict[str, Any] = {
            "groupName": self.group_name,
            "number": self.number,
            "choices": [choice.to_dict() for choice in self.choices],
        }
        return data


@dataclass
class BindingProject:
    project_desc: str
    index_part_no: str
    index_part_desc: str
    required_groups: List[BindingGroup] = field(default_factory=list)

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "BindingProject":
        return cls(
            project_desc=data.get("projectDesc", ""),
            index_part_no=data.get("indexPartNo", ""),
            index_part_desc=data.get("indexPartDesc", ""),
            required_groups=[BindingGroup.from_dict(group) for group in data.get("requiredGroups", [])],
        )

    def to_dict(self) -> Dict[str, Any]:
        return {
            "projectDesc": self.project_desc,
            "indexPartNo": self.index_part_no,
            "indexPartDesc": self.index_part_desc,
            "requiredGroups": [group.to_dict() for group in self.required_groups],
        }


class BindingLibrary:
    def __init__(self, path: Path):
        self.path = path
        self.projects: List[BindingProject] = []

    def load(self) -> None:
        if self.path.exists():
            data = json.loads(self.path.read_text(encoding="utf-8"))
            self.projects = [BindingProject.from_dict(item) for item in data]
        else:
            self.projects = []

    def save(self) -> None:
        payload = [project.to_dict() for project in self.projects]
        self.path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    def export_excel(self, excel_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "绑定料号"
        ws.append(["项目描述", "索引料号", "索引描述", "分组名称", "分组数量", "料号", "描述", "条件模式", "条件料号", "数量"])
        for project in self.projects:
            for group in project.required_groups:
                for choice in group.choices or [BindingChoice(part_no="", desc="")]:
                    ws.append([
                        project.project_desc,
                        project.index_part_no,
                        project.index_part_desc,
                        group.group_name,
                        group.number,
                        choice.part_no,
                        choice.desc,
                        choice.condition_mode or "",
                        ",".join(choice.condition_part_nos),
                        choice.number if choice.number is not None else "",
                    ])
        wb.save(excel_path)

    def import_excel(self, excel_path: Path) -> None:
        wb = load_workbook(excel_path)
        ws = wb.active
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=False))]
        columns = {name: idx for idx, name in enumerate(header)}

        def cell(row_data, key: str, fallback: int | None = None):
            idx = columns.get(key)
            if idx is None:
                idx = fallback
            if idx is None:
                return None
            if idx >= len(row_data):
                return None
            return row_data[idx]

        projects_map: Dict[str, BindingProject] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            row = list(row)
            project_desc = str(cell(row, "项目描述", 0) or "").strip()
            index_part_no = str(cell(row, "索引料号", 1) or "").strip()
            key = f"{project_desc}::{index_part_no}"
            project = projects_map.setdefault(
                key,
                BindingProject(
                    project_desc=project_desc,
                    index_part_no=index_part_no,
                    index_part_desc=str(cell(row, "索引描述", 2) or "").strip(),
                ),
            )
            group_name = str(cell(row, "分组名称", 3) or "").strip()
            if not group_name:
                continue
            group_number_value = cell(row, "分组数量", 4) or 1
            try:
                group_number = float(group_number_value)
            except (TypeError, ValueError):
                group_number = 1.0
            group = _get_or_create_group(project.required_groups, group_name, group_number)
            part_no = str(cell(row, "料号", 5) or "").strip()
            desc = str(cell(row, "描述", 6) or "").strip()
            condition_mode = str(cell(row, "条件模式", 7) or "").strip() or None
            condition_part_nos_raw = cell(row, "条件料号", 8) or ""
            condition_part_nos = [item.strip() for item in str(condition_part_nos_raw).split(",") if item.strip()]
            number_value = cell(row, "数量", 9)
            choice = BindingChoice(
                part_no=part_no,
                desc=desc,
                condition_mode=condition_mode,
                condition_part_nos=condition_part_nos,
                number=_parse_number(number_value),
            )
            if part_no:
                group.choices.append(choice)
        self.projects = list(projects_map.values())
        self.save()

    def find_project(self, part_no: str) -> Optional[BindingProject]:
        for project in self.projects:
            if project.index_part_no == part_no:
                return project
        return None

    def add_project(self, project: BindingProject) -> None:
        self.projects.append(project)
        self.save()

    def remove_project(self, project: BindingProject) -> None:
        self.projects = [item for item in self.projects if item is not project]
        self.save()

    def iter_projects(self) -> Iterable[BindingProject]:
        return iter(self.projects)


def _parse_number(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _get_or_create_group(groups: List[BindingGroup], name: str, number: float) -> BindingGroup:
    for group in groups:
        if group.group_name == name:
            return group
    group = BindingGroup(group_name=name, number=number)
    groups.append(group)
    return group
