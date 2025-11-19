from __future__ import annotations

import sys
import threading
import traceback
import re
import webbrowser
from pathlib import Path
from dataclasses import dataclass
from typing import Callable, Dict
from tkinter import (
    BOTH,
    END,
    LEFT,
    RIGHT,
    Y,
    Button,
    Entry,
    Frame,
    Label,
    Listbox,
    Menu,
    Scrollbar,
    StringVar,
    Text,
    Toplevel,
    Tk,
    filedialog,
    messagebox,
)
from tkinter import ttk

import csv
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageTk

from bomcheck_app.binding_library import BindingChoice, BindingGroup, BindingLibrary, BindingProject
from bomcheck_app.config import AppConfig, load_config, save_config
from bomcheck_app.excel_processor import (
    ExcelProcessor,
    SaveWorkbookError,
    format_quantity_text,
    normalize_part_no,
)
from bomcheck_app.models import ExecutionResult
from bomcheck_app.part_assets import PartAsset, PartAssetStore, open_file
from bomcheck_app.system_parts import (
    SystemPartRecord,
    SystemPartRepository,
    generate_system_part_excel,
)

def _resource_path(relative: str) -> Path:
    base_path = getattr(sys, "_MEIPASS", None)
    if base_path:
        return Path(base_path) / relative
    return Path(__file__).resolve().parent / relative


CONFIG_PATH = _resource_path("config.json")


class Application:
    def __init__(self, root: Tk):
        self.root = root
        self.root.title("料号检测系统")
        self.config_path: Path = CONFIG_PATH
        self.system_part_path: Path | None = None
        self.blocked_applicant_path: Path | None = None
        self.system_part_repository: SystemPartRepository | None = None
        self._system_part_repository_path: Path | None = None
        self.system_part_viewer: SystemPartViewer | None = None
        self.blocked_editor: BlockedApplicantEditor | None = None
        self.binding_editor: BindingEditor | None = None
        self.invalid_part_editor: InvalidPartEditor | None = None
        self.important_editor: ImportantMaterialEditor | None = None
        self.data_file_editor: DataFileEditor | None = None
        self.part_asset_manager: PartAssetManager | None = None
        self.part_asset_store: PartAssetStore | None = None
        self._apply_config(self.config_path)
        self.selected_file: Path | None = None
        self._execution_lock = threading.Lock()
        self._execution_thread: threading.Thread | None = None
        self._build_ui()
        self._refresh_config_entry()

    def _build_ui(self) -> None:
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill=BOTH, expand=True)

        operation_tab = Frame(notebook)
        system_part_tab = Frame(notebook)
        config_tab = Frame(notebook)
        notebook.add(operation_tab, text="执行")
        notebook.add(system_part_tab, text="料号查询")
        notebook.add(config_tab, text="配置")

        config_file_frame = Frame(config_tab)
        config_file_frame.pack(fill=BOTH, padx=10, pady=(10, 0))

        Label(config_file_frame, text="配置文件：").pack(side=LEFT)
        self.config_entry = Entry(config_file_frame, width=50, state="readonly")
        self.config_entry.pack(side=LEFT, padx=5)
        Button(
            config_file_frame,
            text="编辑数据文件",
            command=self._open_data_file_editor,
        ).pack(side=LEFT)
        Button(
            config_file_frame, text="重新加载", command=self._reload_config
        ).pack(side=LEFT, padx=5)

        config_action_frame = Frame(config_tab)
        config_action_frame.pack(fill=BOTH, padx=10, pady=10, anchor="w")
        Button(
            config_action_frame,
            text="编辑失效料号库",
            command=self._open_invalid_part_editor,
        ).pack(side=LEFT)
        Button(
            config_action_frame,
            text="编辑绑定料号",
            command=self._open_binding_editor,
        ).pack(side=LEFT, padx=5)
        Button(
            config_action_frame,
            text="编辑重要物料",
            command=self._open_important_material_editor,
        ).pack(side=LEFT)
        Button(
            config_action_frame,
            text="编辑屏蔽申请人",
            command=self._open_blocked_applicant_editor,
        ).pack(side=LEFT, padx=5)
        Button(
            config_action_frame,
            text="维护料号资源",
            command=self._open_part_asset_manager,
        ).pack(side=LEFT, padx=5)

        system_viewer_frame = Frame(system_part_tab)
        system_viewer_frame.pack(fill=BOTH, expand=True)
        self.system_part_viewer = SystemPartViewer(
            system_viewer_frame,
            self.system_part_path,
            asset_store=self.part_asset_store,
            on_repository_update=self._set_system_part_repository,
        )
        self.system_part_viewer.pack(fill=BOTH, expand=True)

        operation_container = Frame(operation_tab)
        operation_container.pack(fill=BOTH, expand=True)

        file_frame = Frame(operation_container)
        file_frame.pack(fill=BOTH, padx=10, pady=(10, 0))

        Label(file_frame, text="选择BOM Excel文件：").pack(side=LEFT)
        self.file_entry = Entry(file_frame, width=50)
        self.file_entry.pack(side=LEFT, padx=5)
        Button(file_frame, text="浏览", command=self._choose_file).pack(side=LEFT)

        execute_frame = Frame(operation_container)
        execute_frame.pack(fill=BOTH, padx=10, pady=5)
        self.execute_button = Button(execute_frame, text="执行", command=self._execute)
        self.execute_button.pack(side=LEFT)

        result_frame = Frame(operation_container)
        result_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)
        Label(result_frame, text="执行结果：").pack(anchor="w")

        text_container = Frame(result_frame)
        text_container.pack(fill=BOTH, expand=True)
        scrollbar = Scrollbar(text_container)
        scrollbar.pack(side=RIGHT, fill="y")
        self.result_text = Text(text_container, height=15, bg="white", fg="black")
        self.result_text.pack(side=LEFT, fill=BOTH, expand=True)
        self.result_text.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.result_text.yview)

    def _apply_config(self, path: Path) -> None:
        config = load_config(path)
        binding_library = BindingLibrary(config.binding_library)
        binding_library.load()
        processor = ExcelProcessor(config)
        part_asset_store = PartAssetStore(config.part_asset_dir)
        self.config_path = path
        self.config = config
        self.binding_library = binding_library
        self.processor = processor
        self.part_asset_store = part_asset_store
        self.system_part_path = config.system_part_db
        self.system_part_repository = None
        self._system_part_repository_path = None
        self.blocked_applicant_path = config.blocked_applicants
        if self.system_part_viewer is not None:
            self.system_part_viewer.update_path(config.system_part_db)
            self.system_part_viewer.update_asset_store(part_asset_store)
        if self.blocked_editor is not None:
            self.blocked_editor.update_path(config.blocked_applicants)
        self._refresh_config_entry()

    def _refresh_config_entry(self) -> None:
        if hasattr(self, "config_entry"):
            self.config_entry.config(state="normal")
            self.config_entry.delete(0, END)
            self.config_entry.insert(0, str(self.config_path))
            self.config_entry.config(state="readonly")

    def _reload_config(self) -> None:
        try:
            self._apply_config(self.config_path)
        except Exception as exc:  # pragma: no cover - user feedback
            messagebox.showerror("加载失败", f"重新加载配置失败：{exc}")
        else:
            messagebox.showinfo("配置已更新", f"已重新加载：{self.config_path}")

    def _open_data_file_editor(self) -> None:
        if self._reuse_window(self.data_file_editor):
            return
        self.data_file_editor = DataFileEditor(
            self.root,
            self.config,
            self.config_path.parent,
            self._handle_data_file_save,
            on_close=lambda: setattr(self, "data_file_editor", None),
        )

    def _handle_data_file_save(self, new_config: AppConfig) -> None:
        try:
            save_config(self.config_path, new_config)
            self._apply_config(self.config_path)
        except Exception as exc:  # pragma: no cover - user feedback
            messagebox.showerror("保存失败", f"更新配置失败：{exc}")
        else:
            messagebox.showinfo("配置已更新", "数据文件路径已更新。")

    def _choose_file(self) -> None:
        file_path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx"), ("Excel", "*.xlsm")])
        if file_path:
            self.selected_file = Path(file_path)
            self.file_entry.delete(0, END)
            self.file_entry.insert(0, str(file_path))

    def _execute(self) -> None:
        if not self.selected_file:
            messagebox.showerror("错误", "请先选择Excel文件")
            return
        if not self._execution_lock.acquire(blocking=False):
            messagebox.showinfo("处理中", "上一轮执行尚未完成，请稍候再试。")
            return
        self.execute_button.config(state="disabled")
        thread = threading.Thread(target=self._run_execution, daemon=True)
        self._execution_thread = thread
        thread.start()

    def _run_execution(self) -> None:
        try:
            try:
                result = self.processor.execute(self.selected_file, self.binding_library)
            except SaveWorkbookError as error:
                result = error.result
                self._handle_save_error(error)
            except Exception as exc:  # pragma: no cover - runtime safety
                traceback.print_exc()
                self._update_result_box(f"执行失败：{exc}\n{traceback.format_exc()}", success=False)
                return

            summary_lines = self._build_summary_lines(result)
            success = not result.has_missing
            self._update_result_box("\n".join(summary_lines), success=success)
        finally:
            self.root.after(0, self._on_execution_complete)

    def _on_execution_complete(self) -> None:
        if hasattr(self, "execute_button"):
            self.execute_button.config(state="normal")
        if self._execution_lock.locked():
            self._execution_lock.release()
        self._execution_thread = None

    def _update_result_box(self, message: str, success: bool) -> None:
        def update():
            self.result_text.delete(1.0, END)
            self.result_text.insert(END, message)
            self.result_text.configure(bg="#d4edda" if success else "#f8d7da")

        self.root.after(0, update)

    def _handle_save_error(self, error: SaveWorkbookError) -> None:
        decision_event = threading.Event()
        default_extension = error.path.suffix or ".xlsx"

        def prompt() -> None:
            message = (
                f"无法写入文件：{error.path}\n"
                "该文件可能已在其他程序中打开。是否将结果另存为其他文件？"
            )
            save_elsewhere = messagebox.askyesno("文件被占用", message)
            if save_elsewhere:
                new_path = filedialog.asksaveasfilename(
                    title="另存结果",
                    defaultextension=default_extension,
                    filetypes=[("Excel", "*.xlsx"), ("Excel", "*.xlsm")],
                    initialfile=error.path.name,
                )
                if new_path:
                    try:
                        error.workbook.save(new_path)
                        messagebox.showinfo("保存成功", f"结果已保存到：{new_path}")
                    except PermissionError:
                        messagebox.showerror("保存失败", "目标文件正在使用中，未能保存结果。")
                    except Exception as exc:  # pragma: no cover - user feedback
                        messagebox.showerror("保存失败", f"另存失败：{exc}")
            decision_event.set()

        self.root.after(0, prompt)
        decision_event.wait()

    def _open_binding_editor(self) -> None:
        if self._reuse_window(self.binding_editor):
            return
        self.binding_editor = BindingEditor(
            self.root,
            self.binding_library,
            part_lookup=self._lookup_system_part_desc,
            on_close=lambda: setattr(self, "binding_editor", None),
        )

    def _open_invalid_part_editor(self) -> None:
        if self._reuse_window(self.invalid_part_editor):
            return
        self.invalid_part_editor = InvalidPartEditor(
            self.root,
            self.config.invalid_part_db,
            part_lookup=self._lookup_system_part_desc,
            on_close=lambda: setattr(self, "invalid_part_editor", None),
        )

    def _open_important_material_editor(self) -> None:
        if self._reuse_window(self.important_editor):
            return
        self.important_editor = ImportantMaterialEditor(
            self.root,
            self.config.important_materials,
            on_close=lambda: setattr(self, "important_editor", None),
        )

    def _open_blocked_applicant_editor(self) -> None:
        if not self.blocked_applicant_path:
            messagebox.showerror("缺少配置", "请先在数据文件设置中配置屏蔽申请人列表。")
            return
        self.blocked_editor = BlockedApplicantEditor(
            self.root,
            self.blocked_applicant_path,
            on_close=lambda: setattr(self, "blocked_editor", None),
        )

    def _open_part_asset_manager(self) -> None:
        if not self.part_asset_store:
            messagebox.showerror("缺少配置", "请先在数据文件设置中配置料号资源目录。")
            return
        if self._reuse_window(self.part_asset_manager):
            return
        self.part_asset_manager = PartAssetManager(
            self.root,
            self.part_asset_store,
            part_lookup=self._lookup_system_part_desc,
            on_close=lambda: setattr(self, "part_asset_manager", None),
        )

    def _lookup_system_part_desc(self, part_no: str) -> str:
        repository = self._get_system_part_repository()
        if not repository:
            return ""
        record = repository.find(part_no)
        return record.description if record else ""

    def _get_system_part_repository(self) -> SystemPartRepository | None:
        path = self.system_part_path
        if not path:
            return None
        if (
            self.system_part_repository is not None
            and self._system_part_repository_path == path
        ):
            return self.system_part_repository
        try:
            repository = SystemPartRepository(path)
        except Exception:
            return None
        self._set_system_part_repository(repository)
        return repository

    def _reuse_window(self, editor) -> bool:
        try:
            if editor and editor.top.winfo_exists():
                editor.top.deiconify()
                editor.top.lift()
                editor.top.focus_force()
                return True
        except Exception:
            return False
        return False

    def _set_system_part_repository(
        self, repository: SystemPartRepository | None
    ) -> None:
        self.system_part_repository = repository
        self._system_part_repository_path = (
            repository.path if repository is not None else None
        )

    def _build_summary_lines(self, result: ExecutionResult) -> list[str]:
        lines = [
            f"失效料号数量：{format_quantity_text(result.replacement_summary.total_invalid_found)}",
            f"已标记失效料号数量：{format_quantity_text(result.replacement_summary.total_invalid_previously_marked)}",
            f"已替换数量：{format_quantity_text(result.replacement_summary.total_replaced)}",
        ]
        lines.extend(self._summarize_binding_results(result))
        lines.extend(self._summarize_missing_items(result))
        lines.extend(self._summarize_important_hits(result))
        lines.extend(self._summarize_debug_logs(result))
        return lines

    def _summarize_binding_results(self, result: ExecutionResult) -> list[str]:
        binding_group_count = sum(
            len(res.requirement_results) for res in result.binding_results
        )
        lines = [
            "",
            (
                "绑定料号统计：找到 "
                f"{format_quantity_text(len(result.binding_results))} 组项目，"
                f"需求分组 {format_quantity_text(binding_group_count)} 组"
            ),
        ]
        if not result.binding_results:
            lines.append("（未找到匹配的绑定项目）")
            return lines

        for binding_result in result.binding_results:
            project_header = (
                f"- {binding_result.project_desc} ({binding_result.index_part_no})，"
                f"主料数量：{format_quantity_text(binding_result.matched_quantity)}"
            )
            lines.append(project_header)

            for group_result in binding_result.requirement_results:
                group_line = (
                    "  · "
                    + f"{group_result.group_name}：需求 {format_quantity_text(group_result.required_qty)}，"
                    + f"可用 {format_quantity_text(group_result.available_qty)}，"
                    + f"缺少 {format_quantity_text(group_result.missing_qty)}"
                )
                lines.append(group_line)

                if group_result.matched_details:
                    matched_pairs = [
                        f"{part}:{format_quantity_text(qty)}"
                        for part, qty in group_result.matched_details.items()
                    ]
                    lines.append("    满足料号：" + ", ".join(matched_pairs))

                if group_result.missing_choices:
                    lines.append(
                        "    缺少料号：" + ", ".join(group_result.missing_choices)
                    )

        return lines

    def _summarize_missing_items(self, result: ExecutionResult) -> list[str]:
        if not result.missing_items:
            return []

        lines = ["", "缺失物料："]
        for item in result.missing_items:
            lines.append(
                f"- {item.part_no} {item.desc} 缺少 {format_quantity_text(item.missing_qty)}"
            )
        return lines

    def _summarize_important_hits(self, result: ExecutionResult) -> list[str]:
        lines = [
            "",
            f"重要物料统计：找到 {format_quantity_text(len(result.important_hits))} 组",
        ]
        if not result.important_hits:
            lines.append("（无重要物料命中）")
            return lines

        for hit in result.important_hits:
            lines.append(
                f"- {hit.keyword}（{hit.converted_keyword}）：{format_quantity_text(hit.total_quantity)}"
            )
            if hit.matched_parts:
                matched_text = ", ".join(
                    f"{part}:{format_quantity_text(qty)}"
                    for part, qty in hit.matched_parts.items()
                )
                lines.append(f"    命中料号：{matched_text}")
        return lines

    def _summarize_debug_logs(self, result: ExecutionResult) -> list[str]:
        if not result.debug_logs:
            return []

        return ["", "调试信息：", *[f"- {log}" for log in result.debug_logs]]


class DataFileEditor:
    def __init__(
        self,
        master,
        config: AppConfig,
        base_dir: Path,
        on_save: Callable[[AppConfig], None],
        *,
        on_close: Callable[[], None] | None = None,
    ) -> None:
        self.base_dir = base_dir
        self.on_save = on_save
        self.on_close = on_close
        self.top = Toplevel(master)
        self.top.title("数据文件设置")
        self.top.transient(master)
        self.top.grab_set()
        self.top.protocol("WM_DELETE_WINDOW", self._handle_close)

        self.invalid_var = StringVar(value=str(config.invalid_part_db))
        self.binding_var = StringVar(value=str(config.binding_library))
        self.important_var = StringVar(value=str(config.important_materials))
        self.system_part_var = StringVar(value=str(config.system_part_db))
        self.blocked_var = StringVar(value=str(config.blocked_applicants))
        self.asset_var = StringVar(value=str(config.part_asset_dir))

        self._build_ui()

    def _handle_close(self) -> None:
        if self.on_close:
            try:
                self.on_close()
            except Exception:
                pass
        self.top.destroy()

    def _build_ui(self) -> None:
        frame = Frame(self.top)
        frame.pack(fill=BOTH, expand=True, padx=10, pady=10)

        self._build_file_selector(
            frame,
            row=0,
            label_text="失效料号数据库：",
            text_var=self.invalid_var,
            filetypes=[("Excel", "*.xlsx"), ("Excel", "*.xlsm"), ("所有文件", "*.*")],
        )
        self._build_file_selector(
            frame,
            row=1,
            label_text="绑定料号数据库：",
            text_var=self.binding_var,
            filetypes=[("绑定库", "*.js *.json"), ("所有文件", "*.*")],
        )
        self._build_file_selector(
            frame,
            row=2,
            label_text="重要物料清单：",
            text_var=self.important_var,
            filetypes=[("文本", "*.txt"), ("所有文件", "*.*")],
        )
        Label(frame, text="系统料号文件：").grid(row=3, column=0, sticky="w", pady=5)
        Entry(frame, textvariable=self.system_part_var, width=50).grid(
            row=3, column=1, padx=5, sticky="ew"
        )
        Button(
            frame,
            text="浏览",
            command=lambda: self._choose_file(
                self.system_part_var,
                [("系统料号", "*.tsv *.xlsx *.xlsm"), ("所有文件", "*.*")],
            ),
        ).grid(row=3, column=2)
        Button(frame, text="执行", command=self._process_system_parts).grid(
            row=4, column=1, sticky="w", pady=(0, 5)
        )
        self._build_file_selector(
            frame,
            row=5,
            label_text="屏蔽申请人列表：",
            text_var=self.blocked_var,
            filetypes=[("文本", "*.txt"), ("所有文件", "*.*")],
        )
        self._build_file_selector(
            frame,
            row=6,
            label_text="料号资源目录：",
            text_var=self.asset_var,
            filetypes=[("目录", "*.*")],
            is_directory=True,
        )

        button_frame = Frame(self.top)
        button_frame.pack(fill=BOTH, pady=(0, 10))
        Button(button_frame, text="保存", command=self._on_save).pack(side=LEFT, padx=10)
        Button(button_frame, text="取消", command=self.top.destroy).pack(side=LEFT)

    def _build_file_selector(
        self,
        frame: Frame,
        *,
        row: int,
        label_text: str,
        text_var: StringVar,
        filetypes: list[tuple[str, str]],
        is_directory: bool = False,
    ) -> None:
        Label(frame, text=label_text).grid(row=row, column=0, sticky="w", pady=5)
        entry = Entry(frame, textvariable=text_var, width=50)
        entry.grid(row=row, column=1, padx=5, sticky="ew")
        Button(
            frame,
            text="浏览",
            command=lambda var=text_var, types=filetypes, as_dir=is_directory: self._choose_file(
                var, types, as_dir
            ),
        ).grid(row=row, column=2)
        frame.columnconfigure(1, weight=1)

    def _choose_file(
        self, var: StringVar, filetypes: list[tuple[str, str]], is_directory: bool = False
    ) -> None:  # pragma: no cover - user interaction
        if is_directory:
            selected = filedialog.askdirectory()
        else:
            selected = filedialog.askopenfilename(filetypes=filetypes)
        if selected:
            var.set(selected)

    def _on_save(self) -> None:
        invalid_path = self.invalid_var.get().strip()
        binding_path = self.binding_var.get().strip()
        important_path = self.important_var.get().strip()
        system_part_path = self.system_part_var.get().strip()
        blocked_path = self.blocked_var.get().strip()
        asset_path = self.asset_var.get().strip()

        if not invalid_path or not binding_path:
            messagebox.showerror("保存失败", "请完整填写数据库文件路径。")
            return
        if not important_path:
            messagebox.showerror("保存失败", "请填写重要物料清单路径。")
            return
        if not system_part_path:
            messagebox.showerror("保存失败", "请填写系统料号文件路径。")
            return
        if not blocked_path:
            messagebox.showerror("保存失败", "请填写屏蔽申请人列表路径。")
            return
        if not asset_path:
            messagebox.showerror("保存失败", "请填写料号资源目录。")
            return

        new_config = AppConfig(
            invalid_part_db=self._normalize_path(invalid_path),
            binding_library=self._normalize_path(binding_path),
            important_materials=self._normalize_path(important_path),
            system_part_db=self._normalize_path(system_part_path),
            blocked_applicants=self._normalize_path(blocked_path),
            part_asset_dir=self._normalize_path(asset_path),
        )

        try:
            self.on_save(new_config)
        except Exception as exc:  # pragma: no cover - user feedback
            messagebox.showerror("保存失败", f"无法更新配置：{exc}")
            return

        self.top.destroy()

    def _process_system_parts(self) -> None:
        source_path = self.system_part_var.get().strip()
        if not source_path:
            messagebox.showerror("处理失败", "请先选择系统料号原始文件。")
            return

        invalid_path = self.invalid_var.get().strip()
        blocked_path = self.blocked_var.get().strip()
        if not invalid_path:
            messagebox.showerror("处理失败", "请先设置失效料号数据库路径。")
            return
        if not blocked_path:
            messagebox.showerror("处理失败", "请先设置屏蔽申请人列表路径。")
            return

        try:
            source = self._normalize_path(source_path)
            invalid = self._normalize_path(invalid_path)
            blocked = self._normalize_path(blocked_path)
        except Exception as exc:  # pragma: no cover - defensive
            messagebox.showerror("处理失败", f"路径解析失败：{exc}")
            return

        if not source.exists():
            messagebox.showerror("处理失败", f"未找到系统料号原始文件：{source}")
            return
        if not invalid.exists():
            messagebox.showerror("处理失败", f"未找到失效料号数据库：{invalid}")
            return
        blocked.parent.mkdir(parents=True, exist_ok=True)
        if not blocked.exists():
            blocked.touch()

        try:
            output_path = generate_system_part_excel(source, invalid, blocked)
        except Exception as exc:
            messagebox.showerror("处理失败", f"系统料号处理失败：{exc}")
            return

        self.system_part_var.set(str(output_path))
        messagebox.showinfo("完成", f"系统料号Excel已生成：{output_path}")

    def _normalize_path(self, raw_path: str) -> Path:
        path = Path(raw_path)
        if not path.is_absolute():
            path = self.base_dir / path
        return path


class SystemPartViewer(Frame):
    def __init__(
        self,
        master,
        path: Path | None,
        *,
        asset_store: PartAssetStore | None = None,
        on_repository_update: Callable[[SystemPartRepository | None], None] | None = None,
    ) -> None:
        super().__init__(master)
        self.path: Path | None = path
        self.repository: SystemPartRepository | None = None
        self.asset_store = asset_store
        self.search_var = StringVar()
        self.status_var = StringVar()
        self.on_repository_update = on_repository_update
        self._hover_item: str | None = None
        self._hover_coords: tuple[int, int] | None = None
        self._preview_after: str | None = None
        self._preview_window: Toplevel | None = None
        self._preview_photo: ImageTk.PhotoImage | None = None
        self._preview_image_label: Label | None = None
        self._preview_slideshow_after: str | None = None
        self._preview_asset: PartAsset | None = None
        self._preview_image_index: int = 0
        self._preview_hide_after: str | None = None
        self._tree_hover = False
        self._preview_hover = False
        self._preview_base_image: Image.Image | None = None
        self._preview_zoom: float = 1.0
        self._preview_image_frame: Frame | None = None
        self._nav_prev_btn: Button | None = None
        self._nav_next_btn: Button | None = None
        self._image_index_label: Label | None = None
        self._preview_frame_size = (440, 540)
        self._preview_render_size: tuple[int, int] | None = None
        self._preview_image_offset: tuple[int, int] = (0, 0)
        self._drag_start: tuple[int, int] | None = None
        self._active_preview_item: str | None = None
        self._navigation_visible = False
        self._build_ui()
        self.update_path(path)

    def update_path(self, new_path: Path | None) -> None:
        self.path = new_path
        if hasattr(self, "path_label"):
            self.path_label.config(text=str(new_path) if new_path else "未配置")
        if not new_path:
            self.repository = None
            self._clear_tree()
            self.status_var.set("未配置系统料号文件")
            self._notify_repository_update()
            return
        try:
            self.repository = SystemPartRepository(new_path)
        except FileNotFoundError as exc:
            self.repository = None
            self._clear_tree()
            self.status_var.set(str(exc))
            messagebox.showerror("加载失败", str(exc))
            self._notify_repository_update()
            return
        except Exception as exc:  # pragma: no cover - user feedback
            self.repository = None
            self._clear_tree()
            self.status_var.set(f"读取失败：{exc}")
            messagebox.showerror("加载失败", f"读取系统料号失败：{exc}")
            self._notify_repository_update()
            return
        self._refresh_tree()
        self._notify_repository_update()

    def update_asset_store(self, store: PartAssetStore | None) -> None:
        self.asset_store = store
        self._cancel_preview(destroy_window=True)

    def _build_ui(self) -> None:
        main_frame = Frame(self)
        main_frame.pack(fill=BOTH, expand=True)

        path_frame = Frame(main_frame)
        path_frame.pack(fill=BOTH, padx=10, pady=(10, 0))
        Label(path_frame, text="文件路径：").pack(side=LEFT)
        self.path_label = Label(path_frame, text="", anchor="w")
        self.path_label.pack(side=LEFT, fill=BOTH, expand=True)
        Button(path_frame, text="重新加载", command=lambda: self._load_data(show_message=True)).pack(
            side=LEFT, padx=5
        )

        search_frame = Frame(main_frame)
        search_frame.pack(fill=BOTH, padx=10, pady=(5, 0))
        Label(search_frame, text="查询：").pack(side=LEFT)
        search_entry = Entry(search_frame, textvariable=self.search_var)
        search_entry.pack(side=LEFT, fill=BOTH, expand=True)
        search_entry.bind("<Return>", lambda _event: self._perform_search())
        Button(search_frame, text="查找", command=self._perform_search).pack(side=LEFT, padx=5)
        Button(search_frame, text="清除", command=self._clear_search).pack(side=LEFT)
        Button(search_frame, text="导出", command=self._export_records).pack(side=LEFT, padx=(10, 0))

        tree_frame = Frame(main_frame)
        tree_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)
        scrollbar = Scrollbar(tree_frame)
        scrollbar.pack(side=RIGHT, fill=Y)
        self.tree = ttk.Treeview(
            tree_frame,
            columns=("description", "unit", "applicant", "inventory"),
            show="tree headings",
            selectmode="extended",
        )
        self.tree.pack(side=LEFT, fill=BOTH, expand=True)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.tree.yview)
        self.tree.heading("#0", text="分类 / 料号", anchor="w")
        self.tree.heading("description", text="描述", anchor="w")
        self.tree.heading("unit", text="单位", anchor="w")
        self.tree.heading("applicant", text="申请人", anchor="w")
        self.tree.heading("inventory", text="库存", anchor="e")
        self.tree.column("#0", width=240, minwidth=160, anchor="w", stretch=False)
        self.tree.column("description", width=520, minwidth=320, anchor="w", stretch=True)
        self.tree.column("unit", width=70, minwidth=60, anchor="w", stretch=False)
        self.tree.column("applicant", width=180, minwidth=140, anchor="w", stretch=False)
        self.tree.column("inventory", width=90, minwidth=70, anchor="e", stretch=False)
        self.tree.tag_configure("category", font=("TkDefaultFont", 10, "bold"))
        self.tree.tag_configure("category-level-1", background="#eef2ff")
        self.tree.tag_configure("category-level-2", background="#f4f7ff")
        self.tree.tag_configure("category-level-3", background="#f9fbff")
        self.tree.tag_configure("category-level-4", background="#f0fbf0")
        self.tree.tag_configure("part", background="white")

        self.context_menu = Menu(self.tree, tearoff=0)
        self.context_menu.add_command(
            label="复制整行",
            command=lambda: self._copy_selection("row"),
        )
        self.context_menu.add_command(
            label="复制料号",
            command=lambda: self._copy_selection("part"),
        )
        self.context_menu.add_command(
            label="复制描述",
            command=lambda: self._copy_selection("description"),
        )
        self.tree.bind("<Button-3>", self._on_tree_right_click)
        self.tree.bind("<Control-c>", lambda _event: self._copy_selection("row"))
        self.tree.bind("<Command-c>", lambda _event: self._copy_selection("row"))
        self.tree.bind("<Enter>", self._on_tree_enter)
        self.tree.bind("<Motion>", self._on_tree_motion)
        self.tree.bind("<Leave>", self._on_tree_leave)
        self.tree.bind("<Button-1>", self._on_tree_click, add="+")

        status_frame = Frame(main_frame)
        status_frame.pack(fill=BOTH, padx=10, pady=(0, 10))
        Label(status_frame, textvariable=self.status_var, anchor="w").pack(
            side=LEFT, fill=BOTH, expand=True
        )

    def _perform_search(self) -> None:
        self._refresh_tree()

    def _clear_search(self) -> None:
        self.search_var.set("")
        self._refresh_tree()

    def _export_records(self) -> None:
        if not self.repository:
            messagebox.showerror("导出失败", "尚未加载系统料号数据。")
            return
        records = self._get_filtered_records()
        if not records:
            messagebox.showinfo("无数据", "没有可导出的料号记录。")
            return

        file_path = filedialog.asksaveasfilename(
            title="导出料号",
            defaultextension=".xlsx",
            filetypes=[
                ("Excel 工作簿", "*.xlsx"),
                ("CSV 文件", "*.csv"),
                ("所有文件", "*.*"),
            ],
        )
        if not file_path:
            return

        try:
            if file_path.lower().endswith(".csv"):
                self._export_to_csv(file_path, records)
            else:
                self._export_to_excel(file_path, records)
        except Exception as exc:  # pragma: no cover - user feedback
            messagebox.showerror("导出失败", f"导出文件失败：{exc}")
        else:
            messagebox.showinfo("导出成功", f"已导出到：{file_path}")

    def _load_data(self, show_message: bool = False) -> None:
        if not self.path:
            self.status_var.set("未配置系统料号文件")
            self._notify_repository_update()
            return
        try:
            self.repository = SystemPartRepository(self.path)
        except FileNotFoundError as exc:
            self.repository = None
            self._clear_tree()
            self.status_var.set(str(exc))
            messagebox.showerror("加载失败", str(exc))
            self._notify_repository_update()
            return
        except Exception as exc:  # pragma: no cover - user feedback
            self.repository = None
            self._clear_tree()
            self.status_var.set(f"读取失败：{exc}")
            messagebox.showerror("加载失败", f"读取系统料号失败：{exc}")
            self._notify_repository_update()
            return
        self._refresh_tree()
        if show_message:
            messagebox.showinfo("完成", "系统料号已重新加载。")
        self._notify_repository_update()

    def _notify_repository_update(self) -> None:
        if self.on_repository_update:
            self.on_repository_update(self.repository)

    def _refresh_tree(self) -> None:
        if not self.repository:
            self._clear_tree()
            return
        query = self.search_var.get().strip()
        hierarchy = self.repository.build_hierarchy(query or None)
        self._clear_tree()
        self._insert_nodes("", hierarchy)
        total = len(self.repository.records)
        if query:
            matched = len(self.repository.search(query))
            self.status_var.set(f"共 {total} 条，匹配 {matched} 条")
            self._expand_all()
        else:
            self.status_var.set(f"共 {total} 条")

    def _clear_tree(self) -> None:
        for item in self.tree.get_children(""):
            self.tree.delete(item)

    def _on_tree_right_click(self, event) -> str | None:
        item = self.tree.identify_row(event.y)
        if item:
            current_selection = set(self.tree.selection())
            if item not in current_selection:
                self.tree.selection_set(item)
            self.tree.focus(item)
            try:
                self.context_menu.tk_popup(event.x_root, event.y_root)
            finally:
                self.context_menu.grab_release()
        else:
            self.tree.selection_remove(self.tree.selection())
        return "break"

    def _on_tree_enter(self, _event=None) -> None:
        self._tree_hover = True
        self._cancel_preview_hide_timer()

    def _on_tree_motion(self, event) -> None:
        self._tree_hover = True
        self._cancel_preview_hide_timer()
        item = self.tree.identify_row(event.y)
        if item != self._hover_item:
            self._hover_item = item
            self._hover_coords = (event.x_root, event.y_root)
            self._schedule_preview()
        else:
            self._hover_coords = (event.x_root, event.y_root)

    def _on_tree_leave(self, _event) -> None:
        self._hover_item = None
        self._tree_hover = False
        self._schedule_preview_hide()

    def _schedule_preview(self) -> None:
        if self._preview_after:
            try:
                self.after_cancel(self._preview_after)
            except Exception:
                pass
            self._preview_after = None

        if not self._hover_item or not self.asset_store:
            self._schedule_preview_hide()
            return
        if "part" not in self.tree.item(self._hover_item, "tags"):
            self._schedule_preview_hide()
            return

        self._preview_after = self.after(1000, self._show_preview)

    def _schedule_preview_hide(self) -> None:
        self._cancel_preview_hide_timer()
        self._preview_hide_after = self.after(500, self._maybe_close_preview)

    def _maybe_close_preview(self) -> None:
        self._preview_hide_after = None
        if not self._tree_hover and not self._preview_hover:
            self._cancel_preview(destroy_window=True, force=True)

    def _cancel_preview_hide_timer(self) -> None:
        if self._preview_hide_after:
            try:
                self.after_cancel(self._preview_hide_after)
            except Exception:
                pass
            self._preview_hide_after = None

    def _cancel_preview(self, destroy_window: bool = False, force: bool = False) -> None:
        if self._preview_after:
            try:
                self.after_cancel(self._preview_after)
            except Exception:
                pass
            self._preview_after = None
        self._cancel_preview_hide_timer()
        if destroy_window and self._preview_window:
            if force or not self._preview_hover:
                try:
                    self._preview_window.destroy()
                except Exception:
                    pass
                self._preview_window = None
                self._preview_photo = None
                self._preview_image_label = None
                self._preview_image_frame = None
                self._preview_asset = None
                self._preview_image_index = 0
                self._preview_base_image = None
                self._preview_zoom = 1.0
                self._nav_prev_btn = None
                self._nav_next_btn = None
                self._image_index_label = None
                self._preview_render_size = None
                self._preview_image_offset = (0, 0)
                self._drag_start = None
                self._active_preview_item = None
                self._navigation_visible = False
                if self._preview_slideshow_after:
                    try:
                        self.after_cancel(self._preview_slideshow_after)
                    except Exception:
                        pass
                    self._preview_slideshow_after = None
                self._preview_hover = False

    def _on_preview_enter(self, _event=None) -> None:
        self._preview_hover = True
        self._cancel_preview_hide_timer()
        self._pause_slideshow()

    def _on_preview_leave(self, _event=None) -> None:
        self._preview_hover = False
        self._resume_slideshow()
        self._schedule_preview_hide()

    def _show_preview(self) -> None:
        if not self._hover_item or not self.asset_store:
            return
        if "part" not in self.tree.item(self._hover_item, "tags"):
            return
        part_no = self.tree.item(self._hover_item, "text")
        asset = self.asset_store.get(part_no)
        if not asset:
            return
        has_image = bool(asset.images)
        model_path: Path | None = None
        if asset.model_file:
            candidate = self.asset_store.resolve_path(asset.model_file)
            if candidate.exists():
                model_path = candidate
        has_links = bool(asset.remote_links or asset.local_paths or model_path)
        if not has_image and not has_links:
            return
        self._cancel_preview(destroy_window=True, force=True)
        self._preview_window = Toplevel(self)
        self._preview_window.wm_overrideredirect(True)
        self._preview_window.bind("<Enter>", self._on_preview_enter)
        self._preview_window.bind("<Leave>", self._on_preview_leave)
        self._preview_window.bind("<Control-MouseWheel>", self._on_preview_mousewheel)
        self._preview_window.bind("<Control-Button-4>", self._on_preview_mousewheel)
        self._preview_window.bind("<Control-Button-5>", self._on_preview_mousewheel)
        self._position_preview_window()
        self._active_preview_item = self._hover_item

        container = Frame(self._preview_window, bg="white", bd=1, relief="solid")
        container.pack(fill=BOTH, expand=True)
        description = self.tree.item(self._hover_item, "values")[0]
        Label(
            container,
            text=f"{part_no} {description}",
            bg="white",
            anchor="w",
            wraplength=400,
            justify="left",
        ).pack(fill=BOTH, padx=8, pady=(6, 4))

        if asset.images:
            try:
                self._preview_asset = asset
                self._preview_image_index = 0
                self._preview_zoom = 1.0
                self._preview_image_frame = Frame(
                    container, bg="white", width=420, height=320
                )
                self._preview_image_frame.pack(padx=8, pady=4)
                self._preview_image_frame.pack_propagate(False)
                self._preview_image_label = Label(
                    self._preview_image_frame, bg="white"
                )
                self._preview_image_label.place(relx=0.5, rely=0.5, anchor="center")
                self._preview_image_label.bind(
                    "<ButtonPress-1>", self._start_drag_image
                )
                self._preview_image_label.bind(
                    "<B1-Motion>", self._drag_image
                )
                self._preview_image_label.bind(
                    "<ButtonRelease-1>", self._stop_drag_image
                )
                self._preview_image_frame.bind("<Enter>", self._show_navigation)
                self._preview_image_frame.bind("<Leave>", self._hide_navigation)
                self._nav_prev_btn = Button(
                    self._preview_image_frame,
                    text="◀",  # noqa: RUF001 - user-visible arrow
                    width=3,
                    command=self._show_previous_image,
                    bg="#000",
                    fg="white",
                    activebackground="#333",
                    activeforeground="white",
                    relief="flat",
                    bd=0,
                    highlightthickness=0,
                )
                self._nav_next_btn = Button(
                    self._preview_image_frame,
                    text="▶",  # noqa: RUF001 - user-visible arrow
                    width=3,
                    command=self._show_next_image_manual,
                    bg="#000",
                    fg="white",
                    activebackground="#333",
                    activeforeground="white",
                    relief="flat",
                    bd=0,
                    highlightthickness=0,
                )
                self._image_index_label = Label(
                    self._preview_image_frame,
                    text="",
                    bg="#000",
                    fg="white",
                    bd=0,
                    padx=6,
                    pady=2,
                    font=("TkDefaultFont", 9, "bold"),
                )
                self._update_navigation_visibility()
                self._load_preview_image(asset.images[0])
            except Exception:
                Label(container, text="图片预览失败", bg="white", fg="red").pack(
                    fill=BOTH, padx=8, pady=4
                )
            if len(asset.images) > 1:
                self._start_slideshow()

        if model_path or asset.local_paths or asset.remote_links:
            info = Frame(container, bg="white")
            info.pack(fill=BOTH, padx=8, pady=4)
            if model_path:
                Button(
                    info,
                    text="打开3D文件",
                    command=lambda p=model_path: open_file(p),
                ).pack(anchor="w")
            for local in asset.local_paths:
                Button(
                    info,
                    text=f"本地：{local}",
                    command=lambda p=local: open_file(Path(p)),
                    anchor="w",
                ).pack(fill=BOTH, anchor="w")
            for link in asset.remote_links:
                Button(
                    info,
                    text=f"网络：{link}",
                    command=lambda url=link: webbrowser.open(url),
                    anchor="w",
                ).pack(fill=BOTH, anchor="w")

    def _start_slideshow(self) -> None:
        if not self._preview_asset or not self._preview_asset.images:
            return
        if self._preview_hover:
            return
        if self._preview_slideshow_after:
            try:
                self.after_cancel(self._preview_slideshow_after)
            except Exception:
                pass
        self._preview_slideshow_after = self.after(2500, self._show_next_image)

    def _pause_slideshow(self) -> None:
        if self._preview_slideshow_after:
            try:
                self.after_cancel(self._preview_slideshow_after)
            except Exception:
                pass
            self._preview_slideshow_after = None

    def _resume_slideshow(self) -> None:
        if self._preview_hover:
            return
        if self._preview_asset and self._preview_asset.images and len(self._preview_asset.images) > 1:
            self._start_slideshow()

    def _load_preview_image(self, relative_path: str) -> None:
        try:
            self._preview_base_image = self.asset_store.load_image_preview(
                relative_path, max_size=(800, 800)
            )
            self._preview_zoom = self._calculate_fit_zoom()
            self._preview_render_size = None
            self._preview_image_offset = (0, 0)
            self._render_preview_image()
            self._update_navigation_visibility()
        except Exception:
            pass

    def _render_preview_image(self) -> None:
        if not self._preview_base_image:
            return
        if not self._preview_image_label:
            return
        zoom = max(0.1, min(self._preview_zoom, 3.0))
        self._preview_zoom = zoom
        target_width = int(self._preview_base_image.width * zoom)
        target_height = int(self._preview_base_image.height * zoom)
        try:
            resized = self._preview_base_image.resize(
                (max(1, target_width), max(1, target_height)), Image.LANCZOS
            )
            self._preview_photo = ImageTk.PhotoImage(resized)
            self._preview_image_label.config(image=self._preview_photo)
            self._preview_render_size = resized.size
            self._update_image_position()
        except Exception:
            return

    def _show_next_image(self) -> None:
        self._preview_slideshow_after = None
        if not self._preview_window or not self._preview_asset:
            return
        images = self._preview_asset.images
        if len(images) <= 1:
            return
        self._preview_image_index = (self._preview_image_index + 1) % len(images)
        self._preview_zoom = 1.0
        self._preview_image_offset = (0, 0)
        self._load_preview_image(images[self._preview_image_index])
        self._start_slideshow()

    def _show_next_image_manual(self) -> None:
        self._pause_slideshow()
        self._show_next_image()

    def _show_previous_image(self) -> None:
        if not self._preview_asset or not self._preview_asset.images:
            return
        self._pause_slideshow()
        images = self._preview_asset.images
        if len(images) <= 1:
            return
        self._preview_image_index = (self._preview_image_index - 1) % len(images)
        self._preview_zoom = 1.0
        self._preview_image_offset = (0, 0)
        self._load_preview_image(images[self._preview_image_index])

    def _zoom_image(self, delta: float) -> None:
        if not self._preview_base_image:
            return
        self._preview_zoom += delta
        self._render_preview_image()

    def _on_preview_mousewheel(self, event) -> str:
        self._pause_slideshow()
        direction = 0
        if hasattr(event, "delta") and event.delta:
            direction = 1 if event.delta > 0 else -1
        elif getattr(event, "num", None) == 4:
            direction = 1
        elif getattr(event, "num", None) == 5:
            direction = -1
        if direction:
            self._zoom_image(0.1 * direction)
        return "break"

    def _calculate_fit_zoom(self) -> float:
        if not self._preview_base_image:
            return 1.0
        if not self._preview_image_frame:
            return 1.0
        frame_width = self._preview_image_frame.winfo_width() or self._preview_image_frame.winfo_reqwidth()
        frame_height = self._preview_image_frame.winfo_height() or self._preview_image_frame.winfo_reqheight()
        if frame_width <= 1 or frame_height <= 1:
            frame_width, frame_height = 420, 320
        scale = min(frame_width / self._preview_base_image.width, frame_height / self._preview_base_image.height)
        return max(0.1, min(1.0, scale))

    def _constrain_offset(self, offset: tuple[int, int]) -> tuple[int, int]:
        if not self._preview_render_size:
            return offset
        if not self._preview_image_frame:
            return offset
        frame_width = self._preview_image_frame.winfo_width() or self._preview_image_frame.winfo_reqwidth() or 420
        frame_height = self._preview_image_frame.winfo_height() or self._preview_image_frame.winfo_reqheight() or 320
        img_w, img_h = self._preview_render_size
        max_x = max(0, (img_w - frame_width) // 2)
        max_y = max(0, (img_h - frame_height) // 2)
        constrained_x = max(-max_x, min(max_x, offset[0]))
        constrained_y = max(-max_y, min(max_y, offset[1]))
        return constrained_x, constrained_y

    def _update_image_position(self) -> None:
        if not self._preview_image_label or not self._preview_image_frame:
            return
        self._preview_image_offset = self._constrain_offset(self._preview_image_offset)
        self._preview_image_label.place_configure(
            relx=0.5,
            rely=0.5,
            anchor="center",
            x=self._preview_image_offset[0],
            y=self._preview_image_offset[1],
        )

    def _start_drag_image(self, event) -> None:
        self._drag_start = (event.x_root, event.y_root)

    def _drag_image(self, event) -> None:
        if not self._drag_start:
            return
        dx = event.x_root - self._drag_start[0]
        dy = event.y_root - self._drag_start[1]
        self._drag_start = (event.x_root, event.y_root)
        offset_x = self._preview_image_offset[0] + dx
        offset_y = self._preview_image_offset[1] + dy
        self._preview_image_offset = self._constrain_offset((offset_x, offset_y))
        self._update_image_position()

    def _stop_drag_image(self, _event=None) -> None:
        self._drag_start = None

    def _show_navigation(self, _event=None) -> None:
        self._navigation_visible = True
        self._update_navigation_visibility(show_controls=True)

    def _hide_navigation(self, _event=None) -> None:
        self._navigation_visible = False
        self._update_navigation_visibility(show_controls=False)

    def _update_navigation_visibility(self, show_controls: bool | None = None) -> None:
        if show_controls is not None:
            self._navigation_visible = show_controls
        if self._image_index_label and self._preview_asset and self._preview_asset.images:
            total = len(self._preview_asset.images)
            current = self._preview_image_index + 1
            self._image_index_label.config(text=f"{current}/{total}")
            self._image_index_label.place(relx=0.95, rely=0.05, anchor="ne")
        elif self._image_index_label:
            self._image_index_label.place_forget()

        if not self._nav_prev_btn or not self._nav_next_btn:
            return
        for button in (self._nav_prev_btn, self._nav_next_btn):
            button.place_forget()

        if (
            not self._preview_asset
            or not self._preview_asset.images
            or len(self._preview_asset.images) <= 1
            or not self._navigation_visible
        ):
            return

        total = len(self._preview_asset.images)
        if self._preview_image_index > 0:
            self._nav_prev_btn.place(relx=0.04, rely=0.5, anchor="w")
        if self._preview_image_index < total - 1:
            self._nav_next_btn.place(relx=0.96, rely=0.5, anchor="e")

    def _position_preview_window(self) -> None:
        if not self._preview_window or not self._hover_coords:
            return
        width, height = self._preview_frame_size
        offset = 20
        x, y = self._hover_coords
        screen_w = self.winfo_vrootwidth()
        screen_h = self.winfo_vrootheight()
        origin_x = self.winfo_vrootx()
        origin_y = self.winfo_vrooty()

        pos_x = x + offset
        if pos_x + width > origin_x + screen_w:
            pos_x = x - width - offset
            pos_x = max(origin_x, min(pos_x, origin_x + screen_w - width))

        pos_y = y + offset
        if pos_y + height > origin_y + screen_h:
            pos_y = y - height - offset
            pos_y = max(origin_y, min(pos_y, origin_y + screen_h - height))

        self._preview_window.geometry(f"{width}x{height}+{int(pos_x)}+{int(pos_y)}")

    def _on_tree_click(self, _event=None) -> None:
        if not _event:
            return
        item = self.tree.identify_row(_event.y)
        if item:
            self._tree_hover = True
            self._hover_item = item
            self._hover_coords = (_event.x_root, _event.y_root)
            if self._preview_window and self._active_preview_item == item:
                self._cancel_preview_hide_timer()
                return
            self._cancel_preview(destroy_window=True, force=True)
            self._schedule_preview()
        else:
            self._cancel_preview(destroy_window=True, force=True)

    def _insert_nodes(self, parent: str, node: Dict[str, Dict], depth: int = 1) -> None:
        for category, child in self._iter_collapsed_children(node, depth):
            if self._should_skip_category(child):
                for record in child.get("parts", []):
                    self._insert_part(parent, record)
                continue
            tags = ("category", f"category-level-{depth}")
            item_id = self.tree.insert(
                parent, "end", text=category, values=("", "", "", ""), tags=tags
            )
            if depth >= self._max_category_depth:
                for record in self._collect_all_parts(child):
                    self._insert_part(item_id, record)
            else:
                self._insert_nodes(item_id, child, depth + 1)
        if depth <= self._max_category_depth:
            for record in node.get("parts", []):
                self._insert_part(parent, record)

    def _insert_part(self, parent: str, record: SystemPartRecord) -> None:
        self.tree.insert(
            parent,
            "end",
            text=record.part_no,
            values=(
                record.description,
                record.unit,
                record.applicant,
                record.inventory_display,
            ),
            tags=("part",),
        )

    @property
    def _max_category_depth(self) -> int:
        return 4

    def _iter_collapsed_children(self, node: Dict[str, Dict], depth: int) -> list[tuple[str, Dict]]:
        children = node.get("children", {})
        collapsed: list[tuple[str, Dict]] = []
        for category in sorted(children):
            collapsed.append(self._collapse_category_path(category, children[category]))
        return collapsed

    def _collapse_category_path(
        self, label: str, node: Dict[str, Dict]
    ) -> tuple[str, Dict[str, Dict]]:
        current_label = label
        current_node = node
        while (
            not current_node.get("parts")
            and len(current_node.get("children", {})) == 1
        ):
            next_label, next_node = next(iter(current_node["children"].items()))
            current_label = f"{current_label} / {next_label}"
            current_node = next_node
        return current_label, current_node

    def _collect_all_parts(self, node: Dict[str, Dict]) -> list[SystemPartRecord]:
        parts = list(node.get("parts", []))
        for child in node.get("children", {}).values():
            parts.extend(self._collect_all_parts(child))
        return parts

    def _expand_all(self) -> None:
        def expand(item: str) -> None:
            self.tree.item(item, open=True)
            for child in self.tree.get_children(item):
                expand(child)

        for item in self.tree.get_children(""):
            expand(item)

    def _should_skip_category(self, node: Dict[str, Dict]) -> bool:
        return not node.get("children") and len(node.get("parts", [])) == 1

    def _get_filtered_records(self) -> list[SystemPartRecord]:
        if not self.repository:
            return []
        query = self.search_var.get().strip()
        if query:
            return self.repository.search(query)
        return list(self.repository.records)

    def _export_to_excel(self, path: str, records: list[SystemPartRecord]) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "系统料号"
        sheet.append(["料号", "描述", "单位", "申请人", "库存"])
        for record in records:
            sheet.append(
                [
                    record.part_no,
                    record.description,
                    record.unit,
                    record.applicant,
                    record.inventory_display,
                ]
            )
        workbook.save(path)
        workbook.close()

    def _export_to_csv(self, path: str, records: list[SystemPartRecord]) -> None:
        with open(path, "w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(["料号", "描述", "单位", "申请人", "库存"])
            for record in records:
                writer.writerow(
                    [
                        record.part_no,
                        record.description,
                        record.unit,
                        record.applicant,
                        record.inventory_display,
                    ]
                )

    def _copy_selection(self, mode: str) -> None:
        items = [
            item
            for item in self.tree.selection()
            if "part" in self.tree.item(item, "tags")
        ]
        if not items:
            messagebox.showinfo("复制失败", "请先选择要复制的料号。")
            return

        lines: list[str] = []
        for item in items:
            part_no = self.tree.item(item, "text")
            description, unit, applicant, inventory = self.tree.item(item, "values")
            if mode == "part":
                lines.append(part_no)
            elif mode == "description":
                lines.append(description)
            else:
                lines.append(
                    "\t".join(
                        [part_no, description, unit, applicant, inventory]
                    )
                )

        clipboard_text = "\n".join(lines)
        self.tree.clipboard_clear()
        self.tree.clipboard_append(clipboard_text)
        self.status_var.set(f"已复制 {len(lines)} 条记录。")


@dataclass
class InvalidPartEntry:
    invalid_part: str = ""
    invalid_desc: str = ""
    replacement_part: str = ""
    replacement_desc: str = ""


class InvalidPartEditor:
    def __init__(
        self,
        master,
        path: Path,
        *,
        part_lookup: Callable[[str], str] | None = None,
        on_close: Callable[[], None] | None = None,
    ) -> None:
        self.path = path
        self.part_lookup = part_lookup
        self.on_close = on_close
        self.entries: list[InvalidPartEntry] = []
        self.selected_index: int | None = None
        self.filter_var = StringVar()
        self._suspend_events = False
        self.top = Toplevel(master)
        self.top.title("失效料号库编辑")
        self.top.protocol("WM_DELETE_WINDOW", self._handle_close)
        self._build_ui()
        self._load_entries()

    def _build_ui(self) -> None:
        path_frame = Frame(self.top)
        path_frame.pack(fill=BOTH, padx=10, pady=(10, 0))
        Label(path_frame, text="文件路径：").pack(side=LEFT)
        self.path_label = Label(path_frame, text=str(self.path), anchor="w")
        self.path_label.pack(side=LEFT, fill=BOTH, expand=True)

        tree_frame = Frame(self.top)
        tree_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)
        scrollbar = Scrollbar(tree_frame)
        scrollbar.pack(side=RIGHT, fill=Y)
        columns = ("invalid_part", "invalid_desc", "replacement_part", "replacement_desc")
        self.tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show="headings",
            selectmode="browse",
            height=12,
        )
        headings = {
            "invalid_part": "失效料号",
            "invalid_desc": "失效描述",
            "replacement_part": "替换料号",
            "replacement_desc": "替换描述",
        }
        for key, title in headings.items():
            self.tree.heading(key, text=title)
            width = 140 if key.endswith("part") else 260
            self.tree.column(key, width=width, anchor="w", stretch=True)
        self.tree.pack(side=LEFT, fill=BOTH, expand=True)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.tree.yview)
        self.tree.bind("<<TreeviewSelect>>", lambda _event: self._on_tree_select())
        self.tree.bind("<Control-v>", self._handle_paste_shortcut)
        self.tree.bind("<Command-v>", self._handle_paste_shortcut)

        action_frame = Frame(self.top)
        action_frame.pack(fill=BOTH, padx=10, pady=(0, 10))
        Button(action_frame, text="新增", command=self._add_entry).pack(side=LEFT)
        Button(action_frame, text="删除", command=self._delete_entry).pack(side=LEFT, padx=5)
        Button(action_frame, text="复制", command=self._copy_selected_entry).pack(side=LEFT)
        Button(action_frame, text="粘贴", command=self._paste_entries).pack(side=LEFT, padx=5)
        Label(action_frame, text="搜索：").pack(side=LEFT, padx=(10, 0))
        search_entry = Entry(action_frame, textvariable=self.filter_var, width=18)
        search_entry.pack(side=LEFT)
        search_entry.bind("<Return>", lambda _event: self._apply_filter())
        Button(action_frame, text="查找", command=self._apply_filter).pack(side=LEFT, padx=5)
        Button(action_frame, text="清空", command=self._clear_filter).pack(side=LEFT)

        detail_frame = Frame(self.top)
        detail_frame.pack(fill=BOTH, padx=10, pady=(0, 10))
        Label(detail_frame, text="失效料号：").grid(row=0, column=0, sticky="w")
        self.invalid_part_var = StringVar()
        Entry(detail_frame, textvariable=self.invalid_part_var).grid(row=0, column=1, sticky="ew")
        Label(detail_frame, text="失效描述：").grid(row=1, column=0, sticky="w")
        self.invalid_desc_var = StringVar()
        Entry(detail_frame, textvariable=self.invalid_desc_var).grid(row=1, column=1, sticky="ew")
        Label(detail_frame, text="替换料号：").grid(row=2, column=0, sticky="w")
        self.replacement_part_var = StringVar()
        Entry(detail_frame, textvariable=self.replacement_part_var).grid(row=2, column=1, sticky="ew")
        Label(detail_frame, text="替换描述：").grid(row=3, column=0, sticky="w")
        self.replacement_desc_var = StringVar()
        Entry(detail_frame, textvariable=self.replacement_desc_var).grid(row=3, column=1, sticky="ew")
        detail_frame.columnconfigure(1, weight=1)
        self.invalid_part_var.trace_add("write", self._on_invalid_part_change)
        self.invalid_desc_var.trace_add("write", self._on_invalid_desc_edit)
        self.replacement_part_var.trace_add("write", self._on_replacement_part_change)
        self.replacement_desc_var.trace_add("write", self._on_replacement_desc_edit)

        button_frame = Frame(self.top)
        button_frame.pack(fill=BOTH, padx=10, pady=(0, 10))
        Button(button_frame, text="保存", command=self._save_entries).pack(side=LEFT)
        Button(button_frame, text="重新载入", command=self._reload_entries).pack(side=LEFT, padx=5)
        Button(button_frame, text="关闭", command=self._handle_close).pack(side=RIGHT)

    def _handle_paste_shortcut(self, _event) -> str:
        self._paste_entries()
        return "break"

    def _copy_selected_entry(self) -> None:
        self._commit_current_entry()
        if self.selected_index is None or not (0 <= self.selected_index < len(self.entries)):
            messagebox.showinfo("复制", "请先选择需要复制的记录。")
            return
        entry = self.entries[self.selected_index]
        row_text = "\t".join(
            [
                entry.invalid_part,
                entry.invalid_desc,
                entry.replacement_part,
                entry.replacement_desc,
            ]
        )
        try:
            self.top.clipboard_clear()
            self.top.clipboard_append(row_text)
        except Exception as exc:
            messagebox.showerror("复制失败", f"无法写入剪贴板：{exc}")

    def _apply_filter(self) -> None:
        self._refresh_tree()

    def _clear_filter(self) -> None:
        if not self.filter_var.get():
            return
        self.filter_var.set("")
        self._refresh_tree()

    def _get_filtered_indices(self) -> list[int]:
        query = self.filter_var.get().strip().lower()
        if not query:
            return list(range(len(self.entries)))
        tokens = [segment.strip().lower() for segment in re.split(r"[\s,;，；]+", query)]
        tokens = [token for token in tokens if token]
        if not tokens:
            return list(range(len(self.entries)))
        filtered: list[int] = []
        for idx, entry in enumerate(self.entries):
            fields = [
                entry.invalid_part.lower(),
                entry.invalid_desc.lower(),
                entry.replacement_part.lower(),
                entry.replacement_desc.lower(),
            ]
            normalized_invalid = normalize_part_no(entry.invalid_part)
            if normalized_invalid:
                fields.append(normalized_invalid.lower())
            normalized_replacement = normalize_part_no(entry.replacement_part)
            if normalized_replacement:
                fields.append(normalized_replacement.lower())
            haystack = " ".join(fields)
            if all(token in haystack for token in tokens):
                filtered.append(idx)
        return filtered

    def _on_invalid_part_change(self, *_args) -> None:
        if self._suspend_events:
            return
        self._auto_fill_invalid_desc()
        self._commit_current_entry()

    def _on_invalid_desc_edit(self, *_args) -> None:
        if self._suspend_events:
            return
        self._commit_current_entry()

    def _auto_fill_invalid_desc(self) -> None:
        self._auto_fill_description(self.invalid_part_var, self.invalid_desc_var)

    def _on_replacement_part_change(self, *_args) -> None:
        if self._suspend_events:
            return
        self._auto_fill_replacement_desc()
        self._commit_current_entry()

    def _on_replacement_desc_edit(self, *_args) -> None:
        if self._suspend_events:
            return
        self._commit_current_entry()

    def _auto_fill_replacement_desc(self) -> None:
        self._auto_fill_description(
            self.replacement_part_var, self.replacement_desc_var
        )

    def _auto_fill_description(
        self, part_var: StringVar, desc_var: StringVar
    ) -> None:
        if not self.part_lookup:
            return
        part_no = part_var.get().strip()
        if not part_no:
            return
        desc = self.part_lookup(part_no)
        if not desc:
            return
        self._suspend_events = True
        desc_var.set(desc)
        self._suspend_events = False

    def _load_entries(self) -> None:
        self.entries = []
        try:
            if self.path.exists():
                workbook = load_workbook(self.path, data_only=True)
                sheet = workbook.active
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue
                    invalid_part = str(row[0]).strip() if row[0] else ""
                    invalid_desc = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                    replacement_part = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                    replacement_desc = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                    if not invalid_part and not replacement_part and not invalid_desc and not replacement_desc:
                        continue
                    self.entries.append(
                        InvalidPartEntry(
                            invalid_part=invalid_part,
                            invalid_desc=invalid_desc,
                            replacement_part=replacement_part,
                            replacement_desc=replacement_desc,
                        )
                    )
                workbook.close()
        except Exception as exc:
            messagebox.showerror("加载失败", f"读取失效料号失败：{exc}")
            self.entries = []
        self._refresh_tree()
        if self.entries:
            self._select_index(0)
        else:
            self._clear_entry_fields()

    def _refresh_tree(self) -> None:
        selected_ids = self.tree.selection()
        previous_id = selected_ids[0] if selected_ids else None
        filtered_indices = self._get_filtered_indices()
        for item in self.tree.get_children():
            self.tree.delete(item)
        for idx in filtered_indices:
            entry = self.entries[idx]
            values = (
                entry.invalid_part,
                entry.invalid_desc,
                entry.replacement_part,
                entry.replacement_desc,
            )
            self.tree.insert("", "end", iid=str(idx), values=values)
        self.tree.update_idletasks()
        valid_ids = {str(idx) for idx in filtered_indices}
        if previous_id and previous_id in valid_ids:
            self.tree.selection_set(previous_id)
            self.tree.focus(previous_id)
            self.tree.see(previous_id)
        elif filtered_indices:
            first_id = str(filtered_indices[0])
            self.tree.selection_set(first_id)
            self.tree.focus(first_id)
            self.tree.see(first_id)
        else:
            self.tree.selection_set(())
            self.selected_index = None
            self._clear_entry_fields()

    def _on_tree_select(self) -> None:
        if self.selected_index is not None:
            self._commit_current_entry()
        selection = self.tree.selection()
        if not selection:
            self.selected_index = None
            self._clear_entry_fields()
            return
        index = int(selection[0])
        if index >= len(self.entries):
            self.selected_index = None
            self._clear_entry_fields()
            return
        self.selected_index = index
        entry = self.entries[index]
        self._suspend_events = True
        self.invalid_part_var.set(entry.invalid_part)
        self.invalid_desc_var.set(entry.invalid_desc)
        self.replacement_part_var.set(entry.replacement_part)
        self.replacement_desc_var.set(entry.replacement_desc)
        self._suspend_events = False
        if entry.invalid_part and not entry.invalid_desc:
            self._auto_fill_invalid_desc()
        if entry.replacement_part and not entry.replacement_desc:
            self._auto_fill_replacement_desc()
        self._commit_current_entry()

    def _clear_entry_fields(self) -> None:
        self._suspend_events = True
        self.invalid_part_var.set("")
        self.invalid_desc_var.set("")
        self.replacement_part_var.set("")
        self.replacement_desc_var.set("")
        self._suspend_events = False

    def _commit_current_entry(self) -> None:
        if self.selected_index is None:
            return
        if not (0 <= self.selected_index < len(self.entries)):
            return
        entry = self.entries[self.selected_index]
        entry.invalid_part = self.invalid_part_var.get().strip()
        entry.invalid_desc = self.invalid_desc_var.get().strip()
        entry.replacement_part = self.replacement_part_var.get().strip()
        entry.replacement_desc = self.replacement_desc_var.get().strip()
        self._update_tree_item(self.selected_index, entry)

    def _update_tree_item(self, index: int, entry: InvalidPartEntry) -> None:
        item_id = str(index)
        if item_id in self.tree.get_children():
            self.tree.item(
                item_id,
                values=(
                    entry.invalid_part,
                    entry.invalid_desc,
                    entry.replacement_part,
                    entry.replacement_desc,
                ),
            )

    def _add_entry(self) -> None:
        self._commit_current_entry()
        self.entries.append(InvalidPartEntry())
        self._refresh_tree()
        self._select_index(len(self.entries) - 1)

    def _delete_entry(self) -> None:
        if self.selected_index is None:
            return
        if not (0 <= self.selected_index < len(self.entries)):
            return
        del self.entries[self.selected_index]
        self._refresh_tree()
        if self.entries:
            self._select_index(min(self.selected_index, len(self.entries) - 1))
        else:
            self.selected_index = None
            self._clear_entry_fields()

    def _paste_entries(self) -> None:
        self._commit_current_entry()
        try:
            raw_text = self.top.clipboard_get()
        except Exception as exc:
            messagebox.showerror("粘贴失败", f"无法读取剪贴板：{exc}")
            return
        rows: list[InvalidPartEntry] = []
        for line in raw_text.splitlines():
            cells = [cell.strip() for cell in line.split("\t")]
            if not any(cells):
                continue
            invalid_part = cells[0] if len(cells) > 0 else ""
            invalid_desc = cells[1] if len(cells) > 1 else ""
            replacement_part = cells[2] if len(cells) > 2 else ""
            replacement_desc = cells[3] if len(cells) > 3 else ""
            entry = InvalidPartEntry(
                invalid_part=invalid_part,
                invalid_desc=invalid_desc,
                replacement_part=replacement_part,
                replacement_desc=replacement_desc,
            )
            if self.part_lookup and entry.invalid_part and not entry.invalid_desc:
                desc = self.part_lookup(entry.invalid_part)
                if desc:
                    entry.invalid_desc = desc
            if self.part_lookup and entry.replacement_part and not entry.replacement_desc:
                desc = self.part_lookup(entry.replacement_part)
                if desc:
                    entry.replacement_desc = desc
            rows.append(entry)
        if not rows:
            return
        self.entries.extend(rows)
        self._refresh_tree()
        self._select_index(len(self.entries) - len(rows))

    def _select_index(self, index: int) -> None:
        if not self.entries:
            self.selected_index = None
            self._clear_entry_fields()
            return
        index = max(0, min(index, len(self.entries) - 1))
        item_id = str(index)
        if not self.tree.exists(item_id):
            visible = self.tree.get_children()
            if visible:
                first_id = visible[0]
                self.tree.selection_set(first_id)
                self.tree.focus(first_id)
                self.tree.see(first_id)
            else:
                self.tree.selection_set(())
                self.selected_index = None
                self._clear_entry_fields()
            return
        self.tree.selection_set(item_id)
        self.tree.focus(item_id)
        self.tree.see(item_id)

    def _collect_clean_entries(self) -> list[InvalidPartEntry]:
        cleaned: list[InvalidPartEntry] = []
        seen_pairs: set[tuple[str, str]] = set()
        for entry in self.entries:
            invalid_part = entry.invalid_part.strip()
            invalid_desc = entry.invalid_desc.strip()
            replacement_part = entry.replacement_part.strip()
            replacement_desc = entry.replacement_desc.strip()
            if not any([invalid_part, invalid_desc, replacement_part, replacement_desc]):
                continue
            key = (
                normalize_part_no(invalid_part),
                normalize_part_no(replacement_part),
            )
            if key in seen_pairs:
                continue
            seen_pairs.add(key)
            cleaned.append(
                InvalidPartEntry(
                    invalid_part=invalid_part,
                    invalid_desc=invalid_desc,
                    replacement_part=replacement_part,
                    replacement_desc=replacement_desc,
                )
            )
        return cleaned

    def _resolve_replacement_conflicts(
        self, entries: list[InvalidPartEntry]
    ) -> list[InvalidPartEntry]:
        replacement_map: dict[str, set[int]] = {}
        for idx, entry in enumerate(entries):
            replacement_key = normalize_part_no(entry.replacement_part)
            if replacement_key:
                replacement_map.setdefault(replacement_key, set()).add(idx)

        indices_to_remove: set[int] = set()
        for idx, entry in enumerate(entries):
            if idx in indices_to_remove:
                continue
            invalid_key = normalize_part_no(entry.invalid_part)
            if not invalid_key:
                continue
            source_indices = [
                source_idx
                for source_idx in replacement_map.get(invalid_key, set())
                if source_idx != idx
            ]
            if not source_indices:
                continue
            for source_idx in source_indices:
                if source_idx in indices_to_remove:
                    continue
                source_entry = entries[source_idx]
                old_key = normalize_part_no(source_entry.replacement_part)
                source_entry.replacement_part = entry.replacement_part
                source_entry.replacement_desc = entry.replacement_desc
                if old_key:
                    holders = replacement_map.get(old_key)
                    if holders:
                        holders.discard(source_idx)
                        if not holders:
                            replacement_map.pop(old_key, None)
                new_key = normalize_part_no(source_entry.replacement_part)
                if new_key:
                    replacement_map.setdefault(new_key, set()).add(source_idx)
            indices_to_remove.add(idx)

        if not indices_to_remove:
            return entries

        return [entry for i, entry in enumerate(entries) if i not in indices_to_remove]

    def _find_duplicate_invalids(self, entries: list[InvalidPartEntry]) -> list[str]:
        groups: dict[str, list[str]] = {}
        for entry in entries:
            invalid_key = normalize_part_no(entry.invalid_part)
            if not invalid_key:
                continue
            groups.setdefault(invalid_key, []).append(entry.invalid_part or invalid_key)

        duplicates: list[str] = []
        for display_list in groups.values():
            if len(display_list) > 1:
                display_values = sorted({value for value in display_list if value})
                if not display_values:
                    continue
                duplicates.append(", ".join(display_values))
        return duplicates

    def _save_entries(self) -> None:
        self._commit_current_entry()
        cleaned_entries = self._collect_clean_entries()
        cleaned_entries = self._resolve_replacement_conflicts(cleaned_entries)
        duplicate_invalids = self._find_duplicate_invalids(cleaned_entries)
        if duplicate_invalids:
            message = "\n".join(duplicate_invalids)
            messagebox.showerror("保存失败", f"存在重复的失效料号：\n{message}")
            return
        self.entries = cleaned_entries
        self._refresh_tree()
        try:
            self.path.parent.mkdir(parents=True, exist_ok=True)
        except Exception as exc:
            messagebox.showerror("保存失败", f"创建目录失败：{exc}")
            return
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "失效料号"
        sheet.append(["失效料号", "失效描述", "替换料号", "替换描述"])
        for entry in cleaned_entries:
            if not any(
                [
                    entry.invalid_part,
                    entry.invalid_desc,
                    entry.replacement_part,
                    entry.replacement_desc,
                ]
            ):
                continue
            sheet.append(
                [
                    entry.invalid_part,
                    entry.invalid_desc,
                    entry.replacement_part,
                    entry.replacement_desc,
                ]
            )
        try:
            workbook.save(self.path)
        except Exception as exc:
            messagebox.showerror("保存失败", f"写入失效料号失败：{exc}")
            return
        finally:
            workbook.close()
        messagebox.showinfo("保存成功", f"已保存到：{self.path}")

    def _reload_entries(self) -> None:
        self._load_entries()
        messagebox.showinfo("完成", "失效料号已重新加载。")

    def _handle_close(self) -> None:
        if self.on_close:
            try:
                self.on_close()
            except Exception:
                pass
        self.top.destroy()


class BlockedApplicantEditor:
    def __init__(
        self,
        master,
        path: Path,
        *,
        on_close: Callable[[], None] | None = None,
    ) -> None:
        self.path = path
        self.on_close = on_close
        self.top = Toplevel(master)
        self.top.title("屏蔽申请人编辑")
        self._build_ui()
        self.update_path(path)
        self.top.protocol("WM_DELETE_WINDOW", self._handle_close)

    def update_path(self, new_path: Path) -> None:
        self.path = new_path
        if hasattr(self, "path_label"):
            self.path_label.config(text=str(self.path))
        self._load_content()

    def _build_ui(self) -> None:
        path_frame = Frame(self.top)
        path_frame.pack(fill=BOTH, padx=10, pady=(10, 0))
        Label(path_frame, text="文件路径：").pack(side=LEFT)
        self.path_label = Label(path_frame, text=str(self.path), anchor="w")
        self.path_label.pack(side=LEFT, fill=BOTH, expand=True)
        Button(path_frame, text="重新加载", command=self._load_content).pack(side=LEFT, padx=5)

        text_frame = Frame(self.top)
        text_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)
        scrollbar = Scrollbar(text_frame)
        scrollbar.pack(side=RIGHT, fill=Y)
        self.text = Text(text_frame, wrap="none")
        self.text.pack(side=LEFT, fill=BOTH, expand=True)
        self.text.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.text.yview)

        button_frame = Frame(self.top)
        button_frame.pack(fill=BOTH, padx=10, pady=(0, 10))
        Button(button_frame, text="保存", command=self._save_content).pack(side=LEFT)
        Button(button_frame, text="关闭", command=self._handle_close).pack(side=RIGHT)

    def _load_content(self) -> None:
        try:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            if not self.path.exists():
                self.path.touch()
            content = self.path.read_text(encoding="utf-8")
        except Exception as exc:  # pragma: no cover - user feedback
            messagebox.showerror("加载失败", f"读取屏蔽申请人失败：{exc}")
            content = ""
        self.text.delete(1.0, END)
        self.text.insert(END, content)

    def _save_content(self) -> None:
        content = self.text.get("1.0", END)
        if content.endswith("\n"):
            content = content[:-1]
        try:
            self.path.write_text(content, encoding="utf-8")
        except Exception as exc:  # pragma: no cover - user feedback
            messagebox.showerror("保存失败", f"写入屏蔽申请人失败：{exc}")
        else:
            messagebox.showinfo("保存成功", f"已保存到：{self.path}")

    def _handle_close(self) -> None:
        if self.on_close:
            self.on_close()
        self.top.destroy()

class BindingEditor:
    CONDITION_MODE_OPTIONS = ("", "ALL", "ANY", "NOTANY")

    def __init__(
        self,
        master,
        binding_library: BindingLibrary,
        *,
        part_lookup: Callable[[str], str] | None = None,
        on_close: Callable[[], None] | None = None,
    ):
        self.binding_library = binding_library
        self.part_lookup = part_lookup
        self.on_close = on_close
        self.top = Toplevel(master)
        self.top.title("绑定料号编辑")
        self.top.protocol("WM_DELETE_WINDOW", self._handle_close)
        self.projects: list[BindingProject] = []
        self.selected_project_index: int | None = None
        self.selected_group_index: int | None = None
        self.selected_choice_index: int | None = None
        self.project_clipboard: BindingProject | None = None
        self.group_clipboard: BindingGroup | None = None
        self._suspend_part_lookup = False
        self._build_ui()
        self._load_data()

    def _build_ui(self) -> None:
        main_frame = Frame(self.top)
        main_frame.pack(fill=BOTH, expand=True)

        # Project list
        project_frame = Frame(main_frame)
        project_frame.pack(side=LEFT, fill=Y, padx=10, pady=10)
        Label(project_frame, text="项目列表").pack(anchor="w")
        project_list_container = Frame(project_frame)
        project_list_container.pack(fill=Y, expand=True)
        self.project_list = Listbox(
            project_list_container,
            exportselection=False,
            height=15,
            activestyle="none",
            selectmode="browse",
        )
        self.project_list.pack(side=LEFT, fill=Y, expand=True)
        project_scrollbar = Scrollbar(
            project_list_container, orient="vertical", command=self.project_list.yview
        )
        project_scrollbar.pack(side=RIGHT, fill=Y)
        self.project_list.config(yscrollcommand=project_scrollbar.set)
        self.project_list.bind("<<ListboxSelect>>", lambda _event: self._on_project_select())
        Button(project_frame, text="新增项目", command=self._add_project).pack(fill=BOTH, pady=2)
        Button(project_frame, text="删除项目", command=self._remove_project).pack(fill=BOTH, pady=2)
        Button(project_frame, text="上移", command=lambda: self._move_project(-1)).pack(fill=BOTH, pady=2)
        Button(project_frame, text="下移", command=lambda: self._move_project(1)).pack(fill=BOTH, pady=2)
        Button(project_frame, text="复制项目", command=self._copy_project).pack(fill=BOTH, pady=2)
        Button(project_frame, text="粘贴项目", command=self._paste_project).pack(fill=BOTH, pady=2)

        # Detail panel
        detail_frame = Frame(main_frame)
        detail_frame.pack(side=LEFT, fill=BOTH, expand=True, padx=10, pady=10)

        basic_frame = Frame(detail_frame)
        basic_frame.pack(fill=BOTH)
        Label(basic_frame, text="项目描述：").grid(row=0, column=0, sticky="w")
        self.project_desc_var = StringVar()
        Entry(basic_frame, textvariable=self.project_desc_var).grid(row=0, column=1, sticky="ew")
        Label(basic_frame, text="索引料号：").grid(row=1, column=0, sticky="w")
        self.project_index_var = StringVar()
        Entry(basic_frame, textvariable=self.project_index_var).grid(row=1, column=1, sticky="ew")
        Label(basic_frame, text="索引描述：").grid(row=2, column=0, sticky="w")
        self.project_index_desc_var = StringVar()
        Entry(basic_frame, textvariable=self.project_index_desc_var).grid(row=2, column=1, sticky="ew")
        self.project_index_var.trace_add("write", self._on_project_index_change)
        self.project_index_desc_var.trace_add("write", self._on_project_index_desc_edit)
        basic_frame.columnconfigure(1, weight=1)

        # Groups
        group_frame = Frame(detail_frame)
        group_frame.pack(fill=BOTH, expand=True, pady=(10, 0))
        group_left = Frame(group_frame)
        group_left.pack(side=LEFT, fill=Y)
        Label(group_left, text="需求分组").pack(anchor="w")
        group_list_container = Frame(group_left)
        group_list_container.pack(fill=Y, expand=True)
        self.group_list = Listbox(
            group_list_container,
            exportselection=False,
            height=10,
            activestyle="none",
            selectmode="browse",
        )
        self.group_list.pack(side=LEFT, fill=Y, expand=True)
        group_scrollbar = Scrollbar(
            group_list_container, orient="vertical", command=self.group_list.yview
        )
        group_scrollbar.pack(side=RIGHT, fill=Y)
        self.group_list.config(yscrollcommand=group_scrollbar.set)
        self.group_list.bind("<<ListboxSelect>>", lambda _event: self._on_group_select())
        Button(group_left, text="新增分组", command=self._add_group).pack(fill=BOTH, pady=2)
        Button(group_left, text="删除分组", command=self._remove_group).pack(fill=BOTH, pady=2)
        Button(group_left, text="复制分组", command=self._copy_group).pack(fill=BOTH, pady=2)
        Button(group_left, text="粘贴分组", command=self._paste_group).pack(fill=BOTH, pady=2)

        group_detail = Frame(group_frame)
        group_detail.pack(side=LEFT, fill=BOTH, expand=True, padx=10)
        Label(group_detail, text="分组名称：").grid(row=0, column=0, sticky="w")
        self.group_name_var = StringVar()
        Entry(group_detail, textvariable=self.group_name_var).grid(row=0, column=1, sticky="ew")
        Label(group_detail, text="需求数量：").grid(row=1, column=0, sticky="w")
        self.group_number_var = StringVar()
        Entry(group_detail, textvariable=self.group_number_var).grid(row=1, column=1, sticky="ew")
        group_detail.columnconfigure(1, weight=1)

        # Choices table
        choice_frame = Frame(group_detail)
        choice_frame.grid(row=2, column=0, columnspan=2, sticky="nsew", pady=(10, 0))
        group_detail.rowconfigure(2, weight=1)
        Label(choice_frame, text="可选料号").pack(anchor="w")
        columns = ("part_no", "desc", "condition_mode", "condition_part_nos", "number")
        self.choice_tree = ttk.Treeview(choice_frame, columns=columns, show="headings", height=6)
        headings = {
            "part_no": "料号",
            "desc": "描述",
            "condition_mode": "条件模式",
            "condition_part_nos": "条件料号",
            "number": "数量",
        }
        for key, title in headings.items():
            self.choice_tree.heading(key, text=title)
            self.choice_tree.column(key, width=100, anchor="w")
        self.choice_tree.pack(fill=BOTH, expand=True)
        self.choice_tree.bind("<<TreeviewSelect>>", lambda _event: self._on_choice_select())

        choice_edit = Frame(choice_frame)
        choice_edit.pack(fill=BOTH, pady=(5, 0))
        Label(choice_edit, text="料号：").grid(row=0, column=0, sticky="w")
        self.choice_part_var = StringVar()
        Entry(choice_edit, textvariable=self.choice_part_var).grid(row=0, column=1, sticky="ew")
        Label(choice_edit, text="描述：").grid(row=1, column=0, sticky="w")
        self.choice_desc_var = StringVar()
        Entry(choice_edit, textvariable=self.choice_desc_var).grid(row=1, column=1, sticky="ew")
        self.choice_part_var.trace_add("write", self._on_choice_part_change)
        self.choice_desc_var.trace_add("write", self._on_choice_desc_edit)
        Label(choice_edit, text="条件模式：").grid(row=2, column=0, sticky="w")
        self.choice_mode_var = StringVar()
        self.choice_mode_combo = ttk.Combobox(
            choice_edit,
            textvariable=self.choice_mode_var,
            values=self.CONDITION_MODE_OPTIONS,
            state="readonly",
        )
        self.choice_mode_combo.grid(row=2, column=1, sticky="ew")
        self.choice_mode_combo.bind(
            "<<ComboboxSelected>>", lambda _event: self._commit_choice_fields()
        )
        Label(choice_edit, text="条件料号：").grid(row=3, column=0, sticky="w")
        self.choice_condition_var = StringVar()
        Entry(choice_edit, textvariable=self.choice_condition_var).grid(row=3, column=1, sticky="ew")
        Label(choice_edit, text="数量：").grid(row=4, column=0, sticky="w")
        self.choice_number_var = StringVar()
        Entry(choice_edit, textvariable=self.choice_number_var).grid(row=4, column=1, sticky="ew")
        choice_edit.columnconfigure(1, weight=1)

        choice_btn_frame = Frame(choice_frame)
        choice_btn_frame.pack(fill=BOTH, pady=(5, 0))
        Button(choice_btn_frame, text="新增料号", command=self._add_choice).pack(side=LEFT, padx=2)
        Button(choice_btn_frame, text="更新料号", command=self._update_choice).pack(side=LEFT, padx=2)
        Button(choice_btn_frame, text="删除料号", command=self._remove_choice).pack(side=LEFT, padx=2)

        # Bottom action buttons
        button_frame = Frame(self.top)
        button_frame.pack(fill=BOTH, padx=10, pady=5)
        Button(button_frame, text="保存", command=self._save).pack(side=LEFT)
        Button(button_frame, text="重新载入", command=self._load_data).pack(side=LEFT, padx=5)
        Button(button_frame, text="导入Excel", command=self._import_excel).pack(side=LEFT, padx=5)
        Button(button_frame, text="导出Excel", command=self._export_excel).pack(side=LEFT, padx=5)
        Button(button_frame, text="关闭", command=self._handle_close).pack(side=RIGHT)

    def _on_project_index_change(self, *_args) -> None:
        if self._suspend_part_lookup:
            return
        self._auto_fill_project_index_desc()

    def _on_project_index_desc_edit(self, *_args) -> None:
        if self._suspend_part_lookup:
            return

    def _auto_fill_project_index_desc(self) -> None:
        if not self.part_lookup:
            return
        part_no = self.project_index_var.get().strip()
        if not part_no:
            return
        desc = self.part_lookup(part_no)
        if not desc:
            return
        self._suspend_part_lookup = True
        self.project_index_desc_var.set(desc)
        self._suspend_part_lookup = False
        self._commit_project_fields()

    def _on_choice_part_change(self, *_args) -> None:
        if self._suspend_part_lookup:
            return
        self._auto_fill_choice_desc()

    def _on_choice_desc_edit(self, *_args) -> None:
        if self._suspend_part_lookup:
            return

    def _auto_fill_choice_desc(self) -> None:
        if not self.part_lookup:
            return
        part_no = self.choice_part_var.get().strip()
        if not part_no:
            return
        desc = self.part_lookup(part_no)
        if not desc:
            return
        self._suspend_part_lookup = True
        self.choice_desc_var.set(desc)
        self._suspend_part_lookup = False
        self._commit_choice_fields()

    def _load_data(self) -> None:
        self.binding_library.load()
        self.projects = [BindingProject.from_dict(project.to_dict()) for project in self.binding_library.iter_projects()]
        self.selected_project_index = None
        self.selected_group_index = None
        self.selected_choice_index = None
        self._refresh_project_list()
        self._clear_project_fields()
        if self.projects:
            self.project_list.selection_set(0)
            self._ensure_project_visible(0)
            self._on_project_select()

    def _refresh_project_list(self) -> None:
        self.project_list.delete(0, END)
        for project in self.projects:
            display = f"{project.project_desc or '未命名'} ({project.index_part_no or '-'})"
            self.project_list.insert(END, display)

    def _ensure_project_visible(self, index: int) -> None:
        if 0 <= index < self.project_list.size():
            self.project_list.see(index)

    def _clear_project_fields(self) -> None:
        self._suspend_part_lookup = True
        self.project_desc_var.set("")
        self.project_index_var.set("")
        self.project_index_desc_var.set("")
        self.group_list.delete(0, END)
        self.group_name_var.set("")
        self.group_number_var.set("")
        for item in self.choice_tree.get_children():
            self.choice_tree.delete(item)
        self.choice_part_var.set("")
        self.choice_desc_var.set("")
        self.choice_mode_var.set("")
        self.choice_condition_var.set("")
        self.choice_number_var.set("")
        self._suspend_part_lookup = False

    def _on_project_select(self) -> None:
        selection = self.project_list.curselection()
        if self.selected_project_index is not None:
            self._commit_choice_fields()
            self._commit_group_fields()
            self._commit_project_fields()
        if not selection:
            self.selected_project_index = None
            self._clear_project_fields()
            return
        self.selected_project_index = selection[0]
        self._ensure_project_visible(self.selected_project_index)
        project = self.projects[self.selected_project_index]
        self._suspend_part_lookup = True
        self.project_desc_var.set(project.project_desc)
        self.project_index_var.set(project.index_part_no)
        self.project_index_desc_var.set(project.index_part_desc)
        self._suspend_part_lookup = False
        if project.index_part_no and not self.project_index_desc_var.get().strip():
            self._auto_fill_project_index_desc()
        self._refresh_group_list()

    def _commit_project_fields(self) -> None:
        if self.selected_project_index is None:
            return
        project = self.projects[self.selected_project_index]
        project.project_desc = self.project_desc_var.get().strip()
        project.index_part_no = self.project_index_var.get().strip()
        project.index_part_desc = self.project_index_desc_var.get().strip()
        display = f"{project.project_desc or '未命名'} ({project.index_part_no or '-'})"
        if 0 <= self.selected_project_index < self.project_list.size():
            current_selection = self.project_list.curselection()
            preserve_selection = (
                len(current_selection) == 1
                and current_selection[0] == self.selected_project_index
            )
            self.project_list.delete(self.selected_project_index)
            self.project_list.insert(self.selected_project_index, display)
            if preserve_selection:
                self.project_list.selection_clear(0, END)
                self.project_list.selection_set(self.selected_project_index)
                self._ensure_project_visible(self.selected_project_index)

    def _refresh_group_list(self) -> None:
        self.group_list.delete(0, END)
        self.selected_group_index = None
        self.selected_choice_index = None
        for group in self.projects[self.selected_project_index].required_groups:
            display = f"{group.group_name or '未命名'} (需求:{group.number})"
            self.group_list.insert(END, display)
        self.group_name_var.set("")
        self.group_number_var.set("")
        for item in self.choice_tree.get_children():
            self.choice_tree.delete(item)
        if self.projects[self.selected_project_index].required_groups:
            self.group_list.selection_set(0)
            self._ensure_group_visible(0)
            self._on_group_select()

    def _ensure_group_visible(self, index: int) -> None:
        if 0 <= index < self.group_list.size():
            self.group_list.see(index)

    def _on_group_select(self) -> None:
        if self.selected_group_index is not None:
            self._commit_choice_fields()
            self._commit_group_fields()
        selection = self.group_list.curselection()
        if not selection:
            self.selected_group_index = None
            self.group_name_var.set("")
            self.group_number_var.set("")
            self._clear_choice_fields()
            return
        self.selected_group_index = selection[0]
        self._ensure_group_visible(self.selected_group_index)
        group = self.projects[self.selected_project_index].required_groups[self.selected_group_index]
        self.group_name_var.set(group.group_name)
        self.group_number_var.set(str(group.number))
        self._refresh_choice_list(auto_select_first=True)

    def _commit_group_fields(self) -> None:
        if self.selected_project_index is None or self.selected_group_index is None:
            return
        group = self.projects[self.selected_project_index].required_groups[self.selected_group_index]
        group.group_name = self.group_name_var.get().strip()
        try:
            group.number = float(self.group_number_var.get()) if self.group_number_var.get().strip() else 1.0
        except ValueError:
            group.number = 1.0
        display = f"{group.group_name or '未命名'} (需求:{group.number})"
        current_selection = self.group_list.curselection()
        preserve_selection = self.selected_group_index in current_selection
        if 0 <= self.selected_group_index < self.group_list.size():
            self.group_list.delete(self.selected_group_index)
            self.group_list.insert(self.selected_group_index, display)
            if preserve_selection:
                self.group_list.selection_set(self.selected_group_index)
                self._ensure_group_visible(self.selected_group_index)

    def _refresh_choice_list(self, auto_select_first: bool = False) -> None:
        self.choice_tree.selection_remove(self.choice_tree.selection())
        for item in self.choice_tree.get_children():
            self.choice_tree.delete(item)
        group = self.projects[self.selected_project_index].required_groups[self.selected_group_index]
        for idx, choice in enumerate(group.choices):
            self.choice_tree.insert(
                "",
                "end",
                iid=str(idx),
                values=(
                    choice.part_no,
                    choice.desc,
                    choice.condition_mode or "",
                    ",".join(choice.condition_part_nos),
                    choice.number if choice.number is not None else "",
                ),
            )
        self._clear_choice_fields()
        if auto_select_first and group.choices:
            first_id = "0"
            self.choice_tree.focus(first_id)
            self.choice_tree.selection_set(first_id)
            self._on_choice_select()
        self.choice_tree.update_idletasks()

    def _clear_choice_fields(self) -> None:
        self.choice_tree.selection_remove(self.choice_tree.selection())
        self.choice_tree.focus("")
        self._suspend_part_lookup = True
        self.choice_part_var.set("")
        self.choice_desc_var.set("")
        self.choice_mode_combo.configure(values=self.CONDITION_MODE_OPTIONS)
        self._set_choice_mode_value("")
        self.choice_condition_var.set("")
        self.choice_number_var.set("")
        self.selected_choice_index = None
        self._suspend_part_lookup = False

    def _on_choice_select(self) -> None:
        selection = self.choice_tree.selection()
        if self.selected_choice_index is not None:
            self._commit_choice_fields()
        if not selection:
            self.selected_choice_index = None
            self._clear_choice_fields()
            return
        self.selected_choice_index = int(selection[0])
        choice = self.projects[self.selected_project_index].required_groups[self.selected_group_index].choices[
            self.selected_choice_index
        ]
        self._suspend_part_lookup = True
        self.choice_part_var.set(choice.part_no)
        self.choice_desc_var.set(choice.desc)
        self._suspend_part_lookup = False
        if choice.part_no and not self.choice_desc_var.get().strip():
            self._auto_fill_choice_desc()
        self._set_choice_mode_value(choice.condition_mode or "")
        self.choice_condition_var.set(",".join(choice.condition_part_nos))
        self.choice_number_var.set("" if choice.number is None else str(choice.number))

    def _set_choice_mode_value(self, value: str) -> None:
        if not hasattr(self, "choice_mode_combo"):
            self.choice_mode_var.set(value)
            return
        current_values = list(self.choice_mode_combo.cget("values"))
        if value not in current_values:
            current_values.append(value)
            self.choice_mode_combo.configure(values=current_values)
        self.choice_mode_var.set(value)

    def _commit_choice_fields(self) -> None:
        if (
            self.selected_project_index is None
            or self.selected_group_index is None
            or self.selected_choice_index is None
        ):
            return
        group = self.projects[self.selected_project_index].required_groups[self.selected_group_index]
        if self.selected_choice_index >= len(group.choices):
            return
        choice = group.choices[self.selected_choice_index]
        choice.part_no = self.choice_part_var.get().strip()
        choice.desc = self.choice_desc_var.get().strip()
        choice.condition_mode = self.choice_mode_var.get().strip() or None
        condition_raw = self.choice_condition_var.get().strip()
        choice.condition_part_nos = [item.strip() for item in condition_raw.split(",") if item.strip()]
        try:
            choice.number = float(self.choice_number_var.get()) if self.choice_number_var.get().strip() else None
        except ValueError:
            choice.number = None
        values = (
            choice.part_no,
            choice.desc,
            choice.condition_mode or "",
            ",".join(choice.condition_part_nos),
            choice.number if choice.number is not None else "",
        )
        item_id = str(self.selected_choice_index)
        if item_id in self.choice_tree.get_children():
            self.choice_tree.item(item_id, values=values)

    def _add_project(self) -> None:
        self._commit_all()
        new_project = BindingProject(project_desc="新项目", index_part_no="", index_part_desc="", required_groups=[])
        self.projects.append(new_project)
        self._refresh_project_list()
        self.project_list.selection_clear(0, END)
        new_index = len(self.projects) - 1
        self.project_list.selection_set(new_index)
        self._ensure_project_visible(new_index)
        self.project_list.event_generate("<<ListboxSelect>>")

    def _remove_project(self) -> None:
        selection = self.project_list.curselection()
        if not selection:
            return
        index = selection[0]
        del self.projects[index]
        self._refresh_project_list()
        self._clear_project_fields()
        self.selected_project_index = None
        if self.projects:
            new_index = min(index, len(self.projects) - 1)
            self.project_list.selection_set(new_index)
            self._ensure_project_visible(new_index)
            self._on_project_select()

    def _move_project(self, direction: int) -> None:
        selection = self.project_list.curselection()
        if not selection:
            return
        index = selection[0]
        target_index = index + direction
        if target_index < 0 or target_index >= len(self.projects):
            return
        self._commit_all()
        self.projects[index], self.projects[target_index] = (
            self.projects[target_index],
            self.projects[index],
        )
        self._refresh_project_list()
        self.selected_project_index = None
        self.selected_group_index = None
        self.selected_choice_index = None
        self.project_list.selection_clear(0, END)
        self.project_list.selection_set(target_index)
        self._ensure_project_visible(target_index)
        self._on_project_select()

    def _copy_project(self) -> None:
        if self.selected_project_index is None:
            messagebox.showwarning("提示", "请先选择项目")
            return
        self._commit_all()
        project = self.projects[self.selected_project_index]
        self.project_clipboard = BindingProject.from_dict(project.to_dict())

    def _paste_project(self) -> None:
        if self.project_clipboard is None:
            messagebox.showwarning("提示", "请先复制项目")
            return
        self._commit_all()
        new_project = BindingProject.from_dict(self.project_clipboard.to_dict())
        self.projects.append(new_project)
        self._refresh_project_list()
        new_index = len(self.projects) - 1
        self.project_list.selection_clear(0, END)
        self.project_list.selection_set(new_index)
        self._ensure_project_visible(new_index)
        self.project_list.event_generate("<<ListboxSelect>>")

    def _add_group(self) -> None:
        if self.selected_project_index is None:
            messagebox.showwarning("提示", "请先选择项目")
            return
        self._commit_project_fields()
        self._commit_group_fields()
        self._commit_choice_fields()
        group = BindingGroup(group_name="新分组", number=1.0, choices=[])
        self.projects[self.selected_project_index].required_groups.append(group)
        self._refresh_group_list()
        new_index = len(self.projects[self.selected_project_index].required_groups) - 1
        self.group_list.selection_set(new_index)
        self._ensure_group_visible(new_index)
        self.group_list.event_generate("<<ListboxSelect>>")

    def _remove_group(self) -> None:
        if self.selected_project_index is None:
            return
        selection = self.group_list.curselection()
        if not selection:
            return
        index = selection[0]
        del self.projects[self.selected_project_index].required_groups[index]
        self._refresh_group_list()
        groups = self.projects[self.selected_project_index].required_groups
        if groups:
            new_index = min(index, len(groups) - 1)
            self.group_list.selection_set(new_index)
            self._ensure_group_visible(new_index)
            self._on_group_select()

    def _copy_group(self) -> None:
        if self.selected_project_index is None or self.selected_group_index is None:
            messagebox.showwarning("提示", "请先选择分组")
            return
        self._commit_choice_fields()
        self._commit_group_fields()
        group = self.projects[self.selected_project_index].required_groups[self.selected_group_index]
        self.group_clipboard = BindingGroup.from_dict(group.to_dict())

    def _paste_group(self) -> None:
        if self.selected_project_index is None:
            messagebox.showwarning("提示", "请先选择项目")
            return
        if self.group_clipboard is None:
            messagebox.showwarning("提示", "请先复制分组")
            return
        self._commit_group_fields()
        self._commit_choice_fields()
        target_project = self.projects[self.selected_project_index]
        new_group = BindingGroup.from_dict(self.group_clipboard.to_dict())
        target_project.required_groups.append(new_group)
        self._refresh_group_list()
        new_index = len(target_project.required_groups) - 1
        self.group_list.selection_clear(0, END)
        self.group_list.selection_set(new_index)
        self._ensure_group_visible(new_index)
        self.group_list.event_generate("<<ListboxSelect>>")

    def _add_choice(self) -> None:
        if self.selected_project_index is None or self.selected_group_index is None:
            messagebox.showwarning("提示", "请先选择分组")
            return
        self._commit_group_fields()
        self._commit_choice_fields()
        group = self.projects[self.selected_project_index].required_groups[self.selected_group_index]
        group.choices.append(BindingChoice(part_no="", desc=""))
        self._refresh_choice_list()
        new_index = len(group.choices) - 1
        self.choice_tree.selection_set(str(new_index))
        self._on_choice_select()

    def _update_choice(self) -> None:
        if self.selected_choice_index is None:
            return
        current_index = self.selected_choice_index
        self._commit_choice_fields()
        self._refresh_choice_list()
        group = self.projects[self.selected_project_index].required_groups[self.selected_group_index]
        if group.choices:
            target_index = min(current_index, len(group.choices) - 1)
            self.choice_tree.selection_set(str(target_index))
            self._on_choice_select()
        else:
            self._clear_choice_fields()

    def _remove_choice(self) -> None:
        if (
            self.selected_project_index is None
            or self.selected_group_index is None
            or self.selected_choice_index is None
        ):
            return
        group = self.projects[self.selected_project_index].required_groups[self.selected_group_index]
        removed_index = self.selected_choice_index
        if self.selected_choice_index < len(group.choices):
            del group.choices[self.selected_choice_index]
        self._refresh_choice_list()
        if group.choices:
            new_index = min(removed_index, len(group.choices) - 1)
            self.choice_tree.selection_set(str(new_index))
            self._on_choice_select()

    def _commit_all(self) -> None:
        self._commit_choice_fields()
        self._commit_group_fields()
        self._commit_project_fields()

    def _import_excel(self) -> None:
        file_path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx"), ("Excel", "*.xlsm")])
        if not file_path:
            return
        try:
            self.binding_library.import_excel(Path(file_path))
        except Exception as exc:
            messagebox.showerror("错误", f"导入失败：{exc}")
            return
        self._load_data()
        messagebox.showinfo("完成", "导入成功")

    def _export_excel(self) -> None:
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not file_path:
            return
        self._commit_all()
        original_projects = self.binding_library.projects
        self.binding_library.projects = self.projects
        try:
            self.binding_library.export_excel(Path(file_path))
        except Exception as exc:
            messagebox.showerror("错误", f"导出失败：{exc}")
            return
        finally:
            self.binding_library.projects = original_projects
        messagebox.showinfo("完成", "导出成功")

    def _save(self) -> None:
        self._commit_all()
        self.binding_library.projects = self.projects
        try:
            self.binding_library.save()
        except Exception as exc:
            messagebox.showerror("错误", f"保存失败：{exc}")
            return
        messagebox.showinfo("完成", "保存成功")

    def _handle_close(self) -> None:
        if self.on_close:
            try:
                self.on_close()
            except Exception:
                pass
        self.top.destroy()


class PartAssetManager:
    def __init__(
        self,
        master,
        store: PartAssetStore,
        *,
        part_lookup: Callable[[str], str] | None = None,
        on_close: Callable[[], None] | None = None,
    ):
        self.store = store
        self.part_lookup = part_lookup or (lambda _p: "")
        self.on_close = on_close
        self.selected_part: str | None = None
        self.current_asset: PartAsset | None = None

        self.top = Toplevel(master)
        self.top.title("料号资源维护")
        self.top.geometry("980x640")
        self.top.protocol("WM_DELETE_WINDOW", self._handle_close)

        self.part_var = StringVar()
        self.desc_var = StringVar()
        self.model_var = StringVar()

        self._part_validator = self.top.register(self._validate_part_no)

        self._build_ui()
        self._load_assets()

    def _build_ui(self) -> None:
        main = Frame(self.top)
        main.pack(fill=BOTH, expand=True, padx=10, pady=10)

        left = Frame(main)
        left.pack(side=LEFT, fill=Y)
        Label(left, text="已配置料号").pack(anchor="w")
        self.asset_listbox = Listbox(left, width=28)
        self.asset_listbox.pack(fill=Y, expand=True)
        self.asset_listbox.bind("<<ListboxSelect>>", self._on_asset_select)
        list_btns = Frame(left)
        list_btns.pack(fill=BOTH, pady=5)
        Button(list_btns, text="新增", command=self._new_asset).pack(side=LEFT)
        Button(list_btns, text="删除", command=self._delete_asset).pack(side=LEFT, padx=5)
        Button(list_btns, text="刷新", command=self._load_assets).pack(side=LEFT)

        right = Frame(main)
        right.pack(side=LEFT, fill=BOTH, expand=True, padx=(10, 0))

        form = Frame(right)
        form.pack(fill=BOTH, expand=True)
        Label(form, text="料号：").grid(row=0, column=0, sticky="e")
        Entry(
            form,
            textvariable=self.part_var,
            width=18,
            validate="key",
            validatecommand=(self._part_validator, "%P"),
        ).grid(row=0, column=1, sticky="we")
        Label(
            form,
            textvariable=self.desc_var,
            anchor="w",
            wraplength=420,
            justify="left",
        ).grid(row=0, column=2, columnspan=2, sticky="w", padx=6)

        Label(form, text="模型文件：").grid(row=1, column=0, sticky="e", pady=(8, 0))
        Entry(form, textvariable=self.model_var, width=40, state="readonly").grid(row=1, column=1, sticky="w", pady=(8, 0))
        model_btns = Frame(form)
        model_btns.grid(row=1, column=2, sticky="w", pady=(8, 0))
        Button(model_btns, text="选择", command=self._choose_model).pack(side=LEFT)
        Button(model_btns, text="打开", command=self._open_model).pack(side=LEFT, padx=4)
        Button(model_btns, text="清除", command=self._clear_model).pack(side=LEFT)

        Label(form, text="本地地址（多行）：").grid(row=2, column=0, sticky="ne", pady=(8, 0))
        self.local_text = Text(form, height=3, width=40)
        self.local_text.grid(row=2, column=1, sticky="we", pady=(8, 0))
        Label(form, text="网络地址（多行）：").grid(row=3, column=0, sticky="ne", pady=(8, 0))
        self.remote_text = Text(form, height=3, width=40)
        self.remote_text.grid(row=3, column=1, sticky="we", pady=(8, 0))

        Label(form, text="图片列表：").grid(row=4, column=0, sticky="ne", pady=(8, 0))
        img_frame = Frame(form)
        img_frame.grid(row=4, column=1, sticky="w", pady=(8, 0))
        self.image_list = Listbox(img_frame, width=50, height=10)
        self.image_list.pack(side=LEFT, fill=BOTH, expand=True)
        img_btns = Frame(form)
        img_btns.grid(row=4, column=2, sticky="nw", padx=6, pady=(8, 0))
        Button(img_btns, text="添加图片", command=self._add_images).pack(fill=BOTH)
        Button(img_btns, text="删除选中", command=self._remove_image).pack(fill=BOTH, pady=4)
        Button(img_btns, text="打开选中", command=self._open_image).pack(fill=BOTH)
        Button(img_btns, text="上移", command=lambda: self._move_image(-1)).pack(
            fill=BOTH, pady=(6, 2)
        )
        Button(img_btns, text="下移", command=lambda: self._move_image(1)).pack(
            fill=BOTH
        )
        action_bar = Frame(right)
        action_bar.pack(fill=BOTH, pady=10)
        Button(action_bar, text="保存资源", command=self._save_asset).pack(side=LEFT)
        Button(action_bar, text="关闭", command=self._handle_close).pack(side=LEFT, padx=6)

        for idx in range(4):
            form.columnconfigure(idx, weight=1)

    def _load_assets(self) -> None:
        self.asset_listbox.delete(0, END)
        self._asset_index: list[str] = []
        for asset in self.store.list_assets():
            desc = self.part_lookup(asset.part_no)
            display = f"{asset.part_no} - {desc}" if desc else asset.part_no
            self._asset_index.append(asset.part_no)
            self.asset_listbox.insert(END, display)
        if self.selected_part:
            self._select_part(self.selected_part)

    def _select_part(self, part_no: str) -> None:
        if part_no in getattr(self, "_asset_index", []):
            idx = self._asset_index.index(part_no)
            self.asset_listbox.selection_clear(0, END)
            self.asset_listbox.selection_set(idx)
            self.asset_listbox.activate(idx)
        self._load_detail(part_no)

    def _on_asset_select(self, _event=None) -> None:
        selection = self.asset_listbox.curselection()
        if not selection:
            return
        index = selection[0]
        part_no = self._asset_index[index]
        self._load_detail(part_no)

    def _load_detail(self, part_no: str | None) -> None:
        self.selected_part = part_no
        asset = self.store.get(part_no) if part_no else None
        if not asset and part_no:
            asset = PartAsset(part_no=normalize_part_no(part_no) or part_no)
        self.current_asset = asset
        self.part_var.set(part_no or "")
        self.desc_var.set(self.part_lookup(part_no) if part_no else "")
        self.model_var.set(asset.model_file if asset else "")
        self.image_list.delete(0, END)
        for rel in asset.images if asset else []:
            self.image_list.insert(END, rel)
        self.local_text.delete("1.0", END)
        self.remote_text.delete("1.0", END)
        if asset:
            if asset.local_paths:
                self.local_text.insert(END, "\n".join(asset.local_paths))
            if asset.remote_links:
                self.remote_text.insert(END, "\n".join(asset.remote_links))

    def _new_asset(self) -> None:
        self.asset_listbox.selection_clear(0, END)
        self._load_detail("")

    def _delete_asset(self) -> None:
        selection = self.asset_listbox.curselection()
        if not selection:
            return
        index = selection[0]
        part_no = self._asset_index[index]
        if not messagebox.askyesno("确认", f"删除 {part_no} 的资源？"):
            return
        self.store.remove(part_no)
        self._load_assets()
        self._load_detail("")

    def _save_asset(self) -> None:
        part_no = self.part_var.get().strip()
        normalized = normalize_part_no(part_no)
        if not normalized:
            messagebox.showerror("保存失败", "请填写有效的料号。")
            return
        asset = self.store.get(normalized) or PartAsset(part_no=normalized)
        asset.part_no = normalized
        asset.model_file = self.model_var.get().strip() or None
        asset.images = list(self.image_list.get(0, END))
        asset.local_paths = _split_lines(self.local_text.get("1.0", END))
        asset.remote_links = _split_lines(self.remote_text.get("1.0", END))
        self.store.upsert(asset)
        self.selected_part = normalized
        self._load_assets()
        messagebox.showinfo("完成", "资源已保存")

    def _add_images(self) -> None:
        part_no = self._require_part_no()
        if not part_no:
            return
        file_paths = filedialog.askopenfilenames(
            title="选择图片",
            filetypes=[("图片", "*.png *.jpg *.jpeg *.bmp *.gif"), ("所有文件", "*.*")],
        )
        if not file_paths:
            return
        added = self.store.add_images(part_no, [Path(p) for p in file_paths])
        for rel in added:
            self.image_list.insert(END, rel)

    def _remove_image(self) -> None:
        selection = self.image_list.curselection()
        if not selection:
            return
        indices = list(selection)
        existing = list(self.image_list.get(0, END))
        for idx in reversed(indices):
            rel_path = existing.pop(idx)
            self.image_list.delete(idx)
            try:
                self.store.resolve_path(rel_path).unlink(missing_ok=True)
            except Exception:
                pass
        asset = self.store.get(self.part_var.get())
        if asset:
            asset.images = existing
            self.store.upsert(asset)

    def _move_image(self, step: int) -> None:
        selection = self.image_list.curselection()
        if not selection:
            return
        idx = selection[0]
        new_idx = idx + step
        if new_idx < 0 or new_idx >= self.image_list.size():
            return
        value = self.image_list.get(idx)
        self.image_list.delete(idx)
        self.image_list.insert(new_idx, value)
        self.image_list.selection_set(new_idx)

    def _open_image(self) -> None:
        selection = self.image_list.curselection()
        if not selection:
            return
        rel = self.image_list.get(selection[0])
        open_file(self.store.resolve_path(rel))

    def _choose_model(self) -> None:
        part_no = self._require_part_no()
        if not part_no:
            return
        file_path = filedialog.askopenfilename(
            title="选择3D文件",
            filetypes=[("模型", "*.step *.stp *.stl *.obj *.iges *.igs"), ("所有文件", "*.*")],
        )
        if not file_path:
            return
        rel = self.store.set_model_file(part_no, Path(file_path))
        self.model_var.set(rel)

    def _clear_model(self) -> None:
        asset = self.store.get(self.part_var.get())
        if not asset:
            return
        asset.model_file = None
        self.store.upsert(asset)
        self.model_var.set("")

    def _open_model(self) -> None:
        rel = self.model_var.get().strip()
        if not rel:
            return
        open_file(self.store.resolve_path(rel))

    def _validate_part_no(self, value: str) -> bool:
        return len(value) <= 15

    def _require_part_no(self) -> str | None:
        part_no = self.part_var.get().strip()
        normalized = normalize_part_no(part_no)
        if not normalized:
            messagebox.showerror("缺少料号", "请先填写有效的料号后再操作。")
            return None
        self.part_var.set(normalized)
        return normalized

    def _handle_close(self) -> None:
        if self.on_close:
            try:
                self.on_close()
            except Exception:
                pass
        self.top.destroy()


def _split_lines(text: str) -> list[str]:
    return [line.strip() for line in text.splitlines() if line.strip()]


class ImportantMaterialEditor:
    def __init__(
        self,
        master,
        path: Path,
        *,
        on_close: Callable[[], None] | None = None,
    ):
        self.path = path
        self.on_close = on_close
        self.top = Toplevel(master)
        self.top.title("重要物料编辑")
        self.top.protocol("WM_DELETE_WINDOW", self._handle_close)
        self._build_ui()
        self._load_content()

    def _build_ui(self) -> None:
        path_frame = Frame(self.top)
        path_frame.pack(fill=BOTH, padx=10, pady=(10, 0))
        Label(path_frame, text="文件路径：").pack(side=LEFT)
        self.path_label = Label(path_frame, text=str(self.path), anchor="w")
        self.path_label.pack(side=LEFT, fill=BOTH, expand=True)

        text_frame = Frame(self.top)
        text_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)
        scrollbar = Scrollbar(text_frame)
        scrollbar.pack(side=RIGHT, fill=Y)
        self.text = Text(text_frame, wrap="none")
        self.text.pack(side=LEFT, fill=BOTH, expand=True)
        self.text.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.text.yview)

        button_frame = Frame(self.top)
        button_frame.pack(fill=BOTH, padx=10, pady=(0, 10))
        Button(button_frame, text="保存", command=self._save_content).pack(side=LEFT)
        Button(button_frame, text="重新加载", command=self._load_content).pack(side=LEFT, padx=5)
        Button(button_frame, text="关闭", command=self._handle_close).pack(side=RIGHT)

    def _load_content(self) -> None:
        try:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            if not self.path.exists():
                self.path.touch()
            content = self.path.read_text(encoding="utf-8")
        except Exception as exc:  # pragma: no cover - user feedback
            messagebox.showerror("加载失败", f"读取重要物料失败：{exc}")
            content = ""
        self.text.delete(1.0, END)
        self.text.insert(END, content)

    def _save_content(self) -> None:
        content = self.text.get("1.0", END)
        if content.endswith("\n"):
            content = content[:-1]
        try:
            self.path.write_text(content, encoding="utf-8")
        except Exception as exc:  # pragma: no cover - user feedback
            messagebox.showerror("保存失败", f"写入重要物料失败：{exc}")
        else:
            messagebox.showinfo("保存成功", f"重要物料已保存到：{self.path}")

    def _handle_close(self) -> None:
        if self.on_close:
            try:
                self.on_close()
            except Exception:
                pass
        self.top.destroy()


def main() -> None:
    root = Tk()
    Application(root)
    root.mainloop()


if __name__ == "__main__":
    main()
