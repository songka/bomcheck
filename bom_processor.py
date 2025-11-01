"""Core processing logic for the BOM check workflow."""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from opencc import OpenCC
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from binding_library import BindingChoice, BindingLibrary, BindingProject

BLACK_FILL = PatternFill("solid", fgColor="000000")
WHITE_FONT = Font(color="FFFFFF")
SUMMARY_SHEET_NAME = "执行结果"
REMAINING_SHEET_NAME = "剩余物料"


@dataclass
class MissingItem:
    part_no: str
    desc: str
    quantity: float


@dataclass
class GroupResult:
    group_name: str
    required: float
    available: float
    missing: float
    choice_details: List[Dict[str, object]] = field(default_factory=list)


@dataclass
class ProjectResult:
    project_desc: str
    index_part_no: str
    index_quantity: float
    groups: List[GroupResult]

    @property
    def is_ok(self) -> bool:
        return all(group.missing <= 0.000001 for group in self.groups)


@dataclass
class ImportantMaterial:
    part_no: str
    desc: str
    quantity: float


@dataclass
class ProcessResult:
    invalid_count: int
    replaced_count: int
    unreplaced_items: List[MissingItem]
    project_results: List[ProjectResult]
    missing_items: List[MissingItem]
    important_materials: List[ImportantMaterial]
    quantity_column: int
    output_path: Path

    @property
    def is_success(self) -> bool:
        return not self.unreplaced_items and not self.missing_items


class BomProcessor:
    """Execute the entire processing pipeline."""

    def __init__(self, invalid_database: Path, binding_library_path: Path, important_materials_path: Path):
        self.invalid_database = invalid_database
        self.binding_library_path = binding_library_path
        self.important_materials_path = important_materials_path
        self._binding_library = BindingLibrary(binding_library_path)
        self._opencc_s2t = OpenCC("s2t")
        self._opencc_t2s = OpenCC("t2s")

    def process(self, bom_path: Path) -> ProcessResult:
        invalid_map = self._load_invalid_map(self.invalid_database)
        projects = self._binding_library.load()
        important_keywords = self._load_important_keywords(self.important_materials_path)

        workbook = load_workbook(bom_path)
        worksheet = workbook.active
        original_max_col = worksheet.max_column

        replacement_col, replacement_desc_col = self._ensure_replacement_columns(worksheet)

        invalid_hits, replaced_count, unreplaced = self._apply_invalid_replacements(
            worksheet, invalid_map, original_max_col, replacement_col, replacement_desc_col
        )

        quantity_column = self._detect_quantity_column(worksheet)
        bom_data, descriptions = self._collect_bom_data(worksheet, quantity_column, replacement_col, replacement_desc_col)
        projects_result, missing_items, used_counts = self._evaluate_projects(projects, bom_data)
        important_matches = self._find_important_materials(bom_data, descriptions, important_keywords)
        remaining_items = self._calculate_remaining_items(bom_data, used_counts)

        self._write_summary_sheet(
            workbook,
            ProcessResult(
                invalid_count=invalid_hits,
                replaced_count=replaced_count,
                unreplaced_items=unreplaced,
                project_results=projects_result,
                missing_items=missing_items,
                important_materials=important_matches,
                quantity_column=quantity_column,
                output_path=bom_path,
            ),
        )
        self._write_remaining_sheet(workbook, remaining_items)
        workbook.save(bom_path)

        return ProcessResult(
            invalid_count=invalid_hits,
            replaced_count=replaced_count,
            unreplaced_items=unreplaced,
            project_results=projects_result,
            missing_items=missing_items,
            important_materials=important_matches,
            quantity_column=quantity_column,
            output_path=bom_path,
        )

    # ------------------------------------------------------------------
    # Loading helpers

    def _load_invalid_map(self, path: Path) -> Dict[str, Tuple[str, str, str]]:
        from openpyxl import load_workbook as load_invalid_wb

        mapping: Dict[str, Tuple[str, str, str]] = {}
        if not path.exists():
            return mapping
        workbook = load_invalid_wb(path)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2):
            invalid_part = _normalized_part(row[0].value)
            if not invalid_part:
                continue
            invalid_desc = _safe_str(row[1].value)
            replacement_part = _normalized_part(row[2].value)
            replacement_desc = _safe_str(row[3].value)
            mapping[invalid_part] = (invalid_desc, replacement_part, replacement_desc)
        return mapping

    def _load_important_keywords(self, path: Path) -> List[str]:
        if not path.exists():
            return []
        keywords: set[str] = set()
        for line in path.read_text(encoding="utf-8").splitlines():
            keyword = line.strip()
            if not keyword:
                continue
            keywords.add(keyword)
            keywords.add(self._opencc_s2t.convert(keyword))
            keywords.add(self._opencc_t2s.convert(keyword))
        return [k for k in keywords if k]

    # ------------------------------------------------------------------
    # Step 1 - replacement

    def _ensure_replacement_columns(self, worksheet) -> Tuple[int, int]:
        max_col = worksheet.max_column
        replacement_col = max_col + 1
        replacement_desc_col = max_col + 2
        worksheet.cell(1, replacement_col).value = "替换料号"
        worksheet.cell(1, replacement_desc_col).value = "替换描述"
        return replacement_col, replacement_desc_col

    def _apply_invalid_replacements(
        self,
        worksheet,
        invalid_map: Dict[str, Tuple[str, str, str]],
        original_max_col: int,
        replacement_col: int,
        replacement_desc_col: int,
    ) -> Tuple[int, int, List[MissingItem]]:
        invalid_hits = 0
        replaced_count = 0
        unreplaced: List[MissingItem] = []
        for row_idx in range(2, worksheet.max_row + 1):
            original_part = _normalized_part(worksheet.cell(row_idx, 1).value)
            if not original_part or original_part not in invalid_map:
                continue
            invalid_hits += 1
            for col_idx in range(1, original_max_col + 1):
                cell = worksheet.cell(row_idx, col_idx)
                cell.fill = BLACK_FILL
                cell.font = WHITE_FONT
            invalid_desc, replacement_part, replacement_desc = invalid_map[original_part]
            if replacement_part:
                worksheet.cell(row_idx, replacement_col).value = replacement_part
                worksheet.cell(row_idx, replacement_desc_col).value = replacement_desc
                replaced_count += 1
            else:
                unreplaced.append(
                    MissingItem(
                        part_no=original_part,
                        desc=invalid_desc or worksheet.cell(row_idx, 2).value or "",
                        quantity=1.0,
                    )
                )
        return invalid_hits, replaced_count, unreplaced

    # ------------------------------------------------------------------
    # Step 2 - quantity detection & BOM collection

    def _detect_quantity_column(self, worksheet) -> int:
        best_col: Optional[int] = None
        best_decimal_count: Optional[int] = None
        max_row = worksheet.max_row
        for col_idx in range(4, worksheet.max_column + 1):
            numeric = True
            decimal_count = 0
            has_value = False
            for row_idx in range(2, max_row + 1):
                value = worksheet.cell(row_idx, col_idx).value
                if value is None or str(value).strip() == "":
                    continue
                has_value = True
                parsed = _to_float(value)
                if parsed is None:
                    numeric = False
                    break
                if abs(parsed - int(parsed)) > 0:
                    decimal_count += 1
            if not has_value or not numeric:
                continue
            if best_col is None or decimal_count < (best_decimal_count or 0):
                best_col = col_idx
                best_decimal_count = decimal_count
        if best_col is None:
            best_col = 4
        return best_col

    def _collect_bom_data(
        self,
        worksheet,
        quantity_column: int,
        replacement_col: int,
        replacement_desc_col: int,
    ) -> Tuple[Dict[str, Dict[str, object]], Dict[str, str]]:
        bom_data: Dict[str, Dict[str, object]] = {}
        descriptions: Dict[str, str] = {}
        for row_idx in range(2, worksheet.max_row + 1):
            original_part = _normalized_part(worksheet.cell(row_idx, 1).value)
            if not original_part:
                continue
            replacement_part = _normalized_part(worksheet.cell(row_idx, replacement_col).value)
            effective_part = replacement_part or original_part
            desc_cell = worksheet.cell(row_idx, 2).value
            replacement_desc = worksheet.cell(row_idx, replacement_desc_col).value
            desc = _safe_str(replacement_desc) or _safe_str(desc_cell)
            quantity_value = _to_float(worksheet.cell(row_idx, quantity_column).value)
            if quantity_value is None or quantity_value <= 0:
                continue
            entry = bom_data.setdefault(effective_part, {"quantity": 0.0, "desc": desc, "rows": []})
            entry["quantity"] += quantity_value
            entry["desc"] = desc or entry.get("desc", "")
            entry["rows"].append(row_idx)
            if desc:
                descriptions[effective_part] = desc
        return bom_data, descriptions

    # ------------------------------------------------------------------
    # Step 3 & 4 - evaluate binding requirements

    def _evaluate_projects(
        self,
        projects: Sequence[BindingProject],
        bom_data: Dict[str, Dict[str, object]],
    ) -> Tuple[List[ProjectResult], List[MissingItem], Dict[str, float]]:
        results: List[ProjectResult] = []
        missing_items: List[MissingItem] = []
        used_counts: Dict[str, float] = {}
        bom_part_numbers = {part_no for part_no, info in bom_data.items() if info.get("quantity", 0) > 0}
        for project in projects:
            index_info = bom_data.get(project.index_part_no)
            index_quantity = float(index_info.get("quantity", 0)) if index_info else 0.0
            if index_quantity <= 0:
                continue
            group_results: List[GroupResult] = []
            for group in project.required_groups:
                valid_choices = [choice for choice in group.choices if _choice_condition_met(choice, bom_part_numbers)]
                if not valid_choices:
                    missing_items.append(
                        MissingItem(
                            part_no=f"{group.group_name} (无可用料号)",
                            desc=f"{project.project_desc} - {group.group_name}",
                            quantity=float(group.number or 1) * index_quantity,
                        )
                    )
                    group_results.append(
                        GroupResult(
                            group_name=group.group_name,
                            required=float(group.number or 1) * index_quantity,
                            available=0,
                            missing=float(group.number or 1) * index_quantity,
                            choice_details=[],
                        )
                    )
                    continue
                required_per_index = _resolve_group_requirement(group, valid_choices)
                total_required = required_per_index * index_quantity
                choice_details: List[Dict[str, object]] = []
                remaining_required = total_required
                total_available = 0.0
                for choice in valid_choices:
                    available_qty = bom_data.get(choice.part_no, {}).get("quantity", 0.0)
                    used_qty = used_counts.get(choice.part_no, 0.0)
                    usable_qty = max(0.0, available_qty - used_qty)
                    take = min(usable_qty, remaining_required)
                    used_counts[choice.part_no] = used_qty + take
                    total_available += usable_qty
                    remaining_required -= take
                    choice_details.append(
                        {
                            "partNo": choice.part_no,
                            "desc": choice.desc,
                            "available": usable_qty,
                            "used": take,
                            "condition": choice.condition_mode,
                            "requiredPerIndex": choice.number,
                        }
                    )
                shortage = max(0.0, remaining_required)
                if shortage > 0:
                    part_label = "/".join(choice.part_no for choice in valid_choices)
                    desc_label = f"{project.project_desc}-{group.group_name}"
                    missing_items.append(MissingItem(part_no=part_label, desc=desc_label, quantity=shortage))
                group_results.append(
                    GroupResult(
                        group_name=group.group_name,
                        required=total_required,
                        available=total_available,
                        missing=shortage,
                        choice_details=choice_details,
                    )
                )
            results.append(
                ProjectResult(
                    project_desc=project.project_desc,
                    index_part_no=project.index_part_no,
                    index_quantity=index_quantity,
                    groups=group_results,
                )
            )
        return results, missing_items, used_counts

    # ------------------------------------------------------------------
    # Step 5 - important materials

    def _find_important_materials(
        self,
        bom_data: Dict[str, Dict[str, object]],
        descriptions: Dict[str, str],
        keywords: Sequence[str],
    ) -> List[ImportantMaterial]:
        if not keywords:
            return []
        matches: List[ImportantMaterial] = []
        for part_no, info in bom_data.items():
            desc = descriptions.get(part_no, info.get("desc", "")) or ""
            desc_simplified = self._opencc_t2s.convert(desc)
            desc_traditional = self._opencc_s2t.convert(desc)
            for keyword in keywords:
                if keyword in desc or keyword in desc_simplified or keyword in desc_traditional:
                    matches.append(ImportantMaterial(part_no=part_no, desc=desc, quantity=float(info.get("quantity", 0))))
                    break
        return matches

    # ------------------------------------------------------------------
    # Step 6 - workbook summary sheets

    def _write_summary_sheet(self, workbook, result: ProcessResult) -> None:
        if SUMMARY_SHEET_NAME in workbook.sheetnames:
            del workbook[SUMMARY_SHEET_NAME]
        sheet = workbook.create_sheet(SUMMARY_SHEET_NAME)
        row_idx = 1
        sheet.cell(row_idx, 1).value = "失效料号数量"
        sheet.cell(row_idx, 2).value = result.invalid_count
        row_idx += 1
        sheet.cell(row_idx, 1).value = "已替换数量"
        sheet.cell(row_idx, 2).value = result.replaced_count
        row_idx += 2
        sheet.cell(row_idx, 1).value = "未替换料号"
        row_idx += 1
        if not result.unreplaced_items:
            sheet.cell(row_idx, 1).value = "无"
            row_idx += 2
        else:
            for item in result.unreplaced_items:
                sheet.cell(row_idx, 1).value = item.part_no
                sheet.cell(row_idx, 2).value = item.desc
                sheet.cell(row_idx, 3).value = item.quantity
                row_idx += 1
            row_idx += 1
        sheet.cell(row_idx, 1).value = "绑定料号统计"
        row_idx += 1
        for project in result.project_results:
            sheet.cell(row_idx, 1).value = f"{project.project_desc} ({project.index_part_no})"
            sheet.cell(row_idx, 2).value = project.index_quantity
            row_idx += 1
            for group in project.groups:
                status = "OK" if group.missing <= 0.000001 else f"缺少 {group.missing}"
                sheet.cell(row_idx, 1).value = f"  - {group.group_name}"
                sheet.cell(row_idx, 2).value = f"需求 {group.required}"
                sheet.cell(row_idx, 3).value = f"可用 {group.available}"
                sheet.cell(row_idx, 4).value = status
                row_idx += 1
            row_idx += 1
        sheet.cell(row_idx, 1).value = "缺失物料"
        row_idx += 1
        if not result.missing_items:
            sheet.cell(row_idx, 1).value = "无"
            row_idx += 2
        else:
            for item in result.missing_items:
                sheet.cell(row_idx, 1).value = item.part_no
                sheet.cell(row_idx, 2).value = item.desc
                sheet.cell(row_idx, 3).value = item.quantity
                row_idx += 1
            row_idx += 1
        sheet.cell(row_idx, 1).value = "重要物料"
        row_idx += 1
        if not result.important_materials:
            sheet.cell(row_idx, 1).value = "无"
        else:
            for material in result.important_materials:
                sheet.cell(row_idx, 1).value = material.part_no
                sheet.cell(row_idx, 2).value = material.desc
                sheet.cell(row_idx, 3).value = material.quantity
                row_idx += 1

    def _write_remaining_sheet(self, workbook, remaining_items: List[Tuple[str, float, str]]) -> None:
        if REMAINING_SHEET_NAME in workbook.sheetnames:
            del workbook[REMAINING_SHEET_NAME]
        sheet = workbook.create_sheet(REMAINING_SHEET_NAME)
        sheet.cell(1, 1).value = "料号"
        sheet.cell(1, 2).value = "描述"
        sheet.cell(1, 3).value = "剩余数量"
        row_idx = 2
        for part_no, qty, desc in remaining_items:
            sheet.cell(row_idx, 1).value = part_no
            sheet.cell(row_idx, 2).value = desc
            sheet.cell(row_idx, 3).value = qty
            row_idx += 1

    # ------------------------------------------------------------------

    def _calculate_remaining_items(
        self,
        bom_data: Dict[str, Dict[str, object]],
        used_counts: Dict[str, float],
    ) -> List[Tuple[str, float, str]]:
        remaining: List[Tuple[str, float, str]] = []
        for part_no, info in bom_data.items():
            quantity = float(info.get("quantity", 0.0))
            used = used_counts.get(part_no, 0.0)
            leftover = quantity - used
            if leftover > 0.000001:
                remaining.append((part_no, leftover, info.get("desc", "")))
        remaining.sort(key=lambda item: item[0])
        return remaining


def _normalized_part(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text.upper()


def _safe_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_float(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        stripped = str(value).strip().replace(",", "")
        if not stripped:
            return None
        return float(stripped)
    except ValueError:
        return None


def _choice_condition_met(choice: BindingChoice, bom_part_numbers: Iterable[str]) -> bool:
    mode = (choice.condition_mode or "").upper()
    if not mode:
        return True
    condition_parts = [part.upper() for part in choice.condition_part_nos]
    presence = [part in bom_part_numbers for part in condition_parts]
    if mode == "ALL":
        return all(presence)
    if mode == "ANY":
        return any(presence)
    if mode == "NOTANY":
        return not any(presence)
    return True


def _resolve_group_requirement(group, choices: Sequence[BindingChoice]) -> float:
    if group.number is not None:
        return float(group.number)
    if len(choices) == 1 and choices[0].number is not None:
        return float(choices[0].number)
    if choices and any(choice.number for choice in choices):
        numbers = [choice.number for choice in choices if choice.number]
        return float(numbers[0]) if numbers else 1.0
    return 1.0
